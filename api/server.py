import asyncio, json, time, uuid, tempfile, os
from pathlib import Path
from fastapi import FastAPI, Request, Query, UploadFile, File, Form, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse, Response
from fastapi.middleware.cors import CORSMiddleware
import httpx
from markitdown import MarkItDown
from markitdown.converters._html_converter import HtmlConverter
import pandas as pd
import pymupdf4llm
_markitdown = MarkItDown()  # stateless, safe to reuse across requests

# Used only for the custom Excel path below (convert_file bypasses the
# MarkItDown XlsxConverter for .xlsx/.xls so it can (a) filter sheets and
# (b) fix the na_rep='NaN' default — see _convert_excel_sheets()).
_html_converter_excel = HtmlConverter()

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*", "X-Client-ID"])

# ── RAG module (Qdrant-backed retrieval) ──────────────────────
try:
    from rag import rag_router, rag_search_chunks
    app.include_router(rag_router)
    _RAG_AVAILABLE = True
except ImportError as _rag_err:
    print(f"[RAG] rag.py not found or qdrant-client missing — RAG endpoints disabled: {_rag_err}")
    _RAG_AVAILABLE = False
    async def rag_search_chunks(*a, **kw): return []

# ── Long-term memory module (markdown source-of-truth + Qdrant index) ──
try:
    from memory import (
        memory_router, memory_save, memory_update, memory_search_chunks,
        SaveMemoryRequest, UpdateMemoryRequest,
    )
    app.include_router(memory_router)
    _MEMORY_AVAILABLE = True
except ImportError as _mem_err:
    print(f"[MEMORY] memory.py not found or qdrant-client missing — memory endpoints disabled: {_mem_err}")
    _MEMORY_AVAILABLE = False
    async def memory_save(*a, **kw): return {"saved": False, "error": "memory module unavailable"}
    async def memory_update(*a, **kw): return {"saved": False, "error": "memory module unavailable"}
    async def memory_search_chunks(*a, **kw): return []

JOBS_DIR = Path("/app/jobs")
JOBS_DIR.mkdir(exist_ok=True)

_sse_queues: list[tuple[str, asyncio.Queue]] = []  # (client_id, queue)
_running_tasks: dict[str, asyncio.Task] = {}       # job_id → asyncio Task (for cancellation)
_ollama_sem    = asyncio.Semaphore(1)              # only ONE Ollama request active at a time

# Shared, long-lived client for /api/ollama-proxy only. Reused across calls
# so httpx's connection pool can keep the TLS connection to the tunnel
# warm — a fresh `async with httpx.AsyncClient()` per call (as used by the
# job runners below, fine there since jobs are infrequent) means a brand
# new TLS handshake through the Cloudflare tunnel on every single request.
# That's cheap once; it's the dominant cost when listImgModels()/modelCaps
# fire several /api/show calls back-to-back, which is exactly the "models
# take 8-15s to show up" symptom.
_proxy_client = httpx.AsyncClient(timeout=10.0)

# ── Graceful shutdown signal ──────────────────────────────────────────────────
# Set by the lifespan handler on shutdown so SSE generators can exit cleanly
# instead of blocking uvicorn's "Waiting for connections to close" phase.
_shutdown_event = asyncio.Event()

@app.on_event("startup")
async def _on_startup():
    _shutdown_event.clear()

@app.on_event("shutdown")
async def _on_shutdown():
    # Wake all waiting SSE generators so they exit their loops immediately.
    _shutdown_event.set()
    # Give generators one tick to see the event and return.
    await asyncio.sleep(0.1)
    await _proxy_client.aclose()

@app.get("/api/health")
def health(): return {"status": "ok", "mode": "backend", "version": "0.5.1"}


# ── Ollama reachability proxy ──────────────────────────────────────────────
# The frontend calls Ollama directly by default (fast path on the home LAN).
# When that fails — e.g. the browser is outside the LAN and the public
# hostname sits behind Cloudflare Access — ollama.js falls back to this
# route, which forwards the same request from inside the home network,
# exactly like /api/generate and /api/generate-chat already do via the
# `ollama_base` field in their job payloads. GET covers /api/tags, /api/ps,
# /api/version; POST covers /api/show. `base` is the same OLLAMA_BASE/
# IMG_API_BASE value the frontend already has configured — this endpoint
# does not widen what a caller could already reach via job submission.
@app.api_route("/api/ollama-proxy/{path:path}", methods=["GET", "POST"])
async def ollama_proxy(path: str, request: Request, base: str = Query(...)):
    target = f"{base.rstrip('/')}/{path}"
    body = await request.body()
    # /api/generate (native image gen) and /v1/images/generations
    # (OpenAI-compat) can legitimately run for a long time — give those a
    # much longer timeout than quick metadata calls like /api/tags or
    # /api/show, which should fail fast if something's actually wrong.
    # Tightened from 10.0s to match ollama.js's new 4500ms client-side
    # ceiling for these same calls — no point holding this connection open
    # past the point the browser has already given up and moved on.
    timeout = 180.0 if 'generat' in path else 5.0
    try:
        resp = await _proxy_client.request(
            request.method,
            target,
            content=body or None,
            headers={"Content-Type": request.headers.get("content-type", "application/json")},
            timeout=timeout,
        )
    except httpx.RequestError as e:
        return JSONResponse({"error": str(e)}, status_code=502)
    return Response(content=resp.content, status_code=resp.status_code,
                     media_type=resp.headers.get("content-type"))


_EXCEL_EXTS = {".xlsx", ".xls"}

def _excel_engine(suffix: str) -> str:
    return "xlrd" if suffix == ".xls" else "openpyxl"


def _peek_excel_sheets(tmp_path: str, suffix: str) -> list[dict]:
    """Return [{name, rows, cols}, ...] for every sheet WITHOUT loading cell
    values — just enough for the frontend's sheet-picker modal. Cheap even
    on large workbooks since it reads sheet dimensions, not cell contents."""
    if suffix == ".xls":
        import xlrd
        book = xlrd.open_workbook(tmp_path, on_demand=True)
        try:
            return [
                {"name": name, "rows": book.sheet_by_name(name).nrows,
                 "cols": book.sheet_by_name(name).ncols}
                for name in book.sheet_names()
            ]
        finally:
            book.release_resources()
    import openpyxl
    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
    try:
        return [
            {"name": ws.title, "rows": ws.max_row or 0, "cols": ws.max_column or 0}
            for ws in wb.worksheets
        ]
    finally:
        wb.close()


def _convert_excel_sheets(tmp_path: str, suffix: str, sheet_names: list[str] | None) -> str:
    """Excel → Markdown, mirroring MarkItDown's XlsxConverter (one '## Sheet'
    heading + table per sheet) but with two fixes:
      1. na_rep='' so empty/dropdown-source cells don't cost tokens as the
         literal string 'NaN'.
      2. Optional `sheet_names` filter so hidden/helper sheets (e.g. ones
         only used to populate dropdown validation lists) can be excluded
         before the content ever reaches the model.
    """
    sheets = pd.read_excel(tmp_path, sheet_name=None, engine=_excel_engine(suffix))
    md_parts = []
    for name, df in sheets.items():
        if sheet_names is not None and name not in sheet_names:
            continue
        html = df.to_html(index=False, na_rep="")
        table_md = _html_converter_excel.convert_string(html).markdown.strip()
        md_parts.append(f"## {name}\n{table_md}")
    return "\n\n".join(md_parts).strip()


@app.post("/api/xlsx-sheets")
async def xlsx_sheets(file: UploadFile = File(...)):
    """Peek at sheet names/sizes in an uploaded Excel file, without doing a
    full conversion. The frontend calls this first so it can offer a sheet
    picker before the real /api/convert-file call."""
    suffix = Path(file.filename).suffix.lower() if file.filename else ""
    if suffix not in _EXCEL_EXTS:
        return JSONResponse({"error": "not an Excel file"}, status_code=400)
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await file.read())
            tmp_path = tmp.name
        sheets = _peek_excel_sheets(tmp_path, suffix)
        return {"filename": file.filename, "sheets": sheets}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


@app.post("/api/convert-file")
async def convert_file(file: UploadFile = File(...), sheets: str | None = Form(None)):
    """Convert an uploaded file to Markdown.
    PDF → pymupdf4llm (layout-aware: tables, headings, columns).
    XLSX/XLS → custom pandas path (sheet filtering + na_rep='' instead of
    MarkItDown's default, which writes empty cells as the literal "NaN").
    Everything else → MarkItDown (docx, pptx, csv, html, xml, json, zip).

    `sheets` (optional): JSON-encoded list of sheet names to include. Only
    used for .xlsx/.xls; ignored otherwise. Omit/None = all sheets (back-
    compat with callers that don't know about sheet selection)."""
    suffix = Path(file.filename).suffix.lower() if file.filename else ""
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await file.read())
            tmp_path = tmp.name
        if suffix == ".pdf":
            markdown = pymupdf4llm.to_markdown(tmp_path)
        elif suffix in _EXCEL_EXTS:
            sheet_filter = json.loads(sheets) if sheets else None
            markdown = _convert_excel_sheets(tmp_path, suffix, sheet_filter)
        else:
            result = _markitdown.convert(tmp_path)
            markdown = result.text_content
        return {"filename": file.filename, "markdown": markdown}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


def job_path(job_id): return JOBS_DIR / f"{job_id}.json"
def reply_path(job_id): return JOBS_DIR / f"{job_id}.reply"   # lightweight per-token accumulator

# ── Jarvis file extraction ────────────────────────────────────────────────────
# Files are saved under /app/jobs/files/<job_id>/ and are NEVER auto-deleted.
JARVIS_FILES_DIR = JOBS_DIR / "files"
JARVIS_FILES_DIR.mkdir(exist_ok=True)

import re as _re

# Matches only the opening tag to extract the filename attribute.
# Single or double quotes are both accepted.  The content is extracted
# manually below so we never hit regex backtracking on large HTML files.
_JARVIS_OPEN_RE = _re.compile(
    r'<jarvis_file\s+name=["\']([^"\']+)["\']\s*>',
)
_JARVIS_CLOSE = '</jarvis_file>'

def extract_jarvis_files(reply: str) -> list[tuple[str, str]]:
    """Return [(filename, content), ...] found in the assistant reply.

    Uses a linear two-pointer scan instead of a greedy/lazy regex so that
    arbitrarily large file contents (HTML pages with thousands of tags,
    minified JS, etc.) are always extracted correctly without backtracking.
    """
    results = []
    pos = 0
    length = len(reply)

    while pos < length:
        # Find the next opening tag from the current position
        m = _JARVIS_OPEN_RE.search(reply, pos)
        if m is None:
            break  # No more file blocks

        filename = m.group(1).strip()
        content_start = m.end()  # byte after the closing '>' of the open tag

        # Find the matching close tag — plain string search, O(n), no backtracking
        close_idx = reply.find(_JARVIS_CLOSE, content_start)
        if close_idx == -1:
            # Truncated / incomplete block — skip (streaming not yet complete)
            break

        content = reply[content_start:close_idx]
        results.append((filename, content))

        # Advance past the close tag for the next iteration
        pos = close_idx + len(_JARVIS_CLOSE)

    return results

def save_jarvis_files(job_id: str, reply: str) -> list[str]:
    """Parse jarvis_file tags from *reply*, write each to /app/jobs/files/<job_id>/.
    Returns list of saved paths (as strings). Safe to call multiple times — overwrites."""
    files = extract_jarvis_files(reply)
    if not files:
        return []
    job_files_dir = JARVIS_FILES_DIR / job_id
    job_files_dir.mkdir(parents=True, exist_ok=True)
    saved = []
    for name, content in files:
        # Sanitise filename: strip leading slashes / path traversal
        safe_name = Path(name).name or "unnamed_file"
        dest = job_files_dir / safe_name
        dest.write_text(content, encoding="utf-8")
        saved.append(str(dest))
    return saved
def read_job(job_id):
    p = job_path(job_id)
    return json.loads(p.read_text()) if p.exists() else None
def write_job(job): job_path(job["id"]).write_text(json.dumps(job))

# Write ONLY the streaming-reply accumulator without touching the full job JSON.
# Called on every chunk — avoids rewriting megabytes of base64 image data per token.
def write_reply_fast(job_id: str, reply: str, thinking: str = ""):
    try:
        reply_path(job_id).write_text(json.dumps({"reply": reply, "thinking": thinking}))
    except Exception:
        pass

# Read back the accumulated reply (used by the reconnect endpoint).
def read_reply_fast(job_id: str) -> tuple[str, str]:
    p = reply_path(job_id)
    if not p.exists():
        return "", ""
    try:
        d = json.loads(p.read_text())
        return d.get("reply", ""), d.get("thinking", "")
    except Exception:
        return "", ""

# Clean up the fast-reply sidecar when the job is fully persisted or deleted.
def _cleanup_reply_file(job_id: str):
    try:
        reply_path(job_id).unlink(missing_ok=True)
    except Exception:
        pass

def get_client_id(req: Request) -> str:
    return req.headers.get("x-client-id", "anonymous")

async def notify(event: dict):
    target = event.get("client_id")
    for cid, q in list(_sse_queues):
        if not target or cid == target:
            await q.put(event)




# ── Image generation job submission ──────────────────────────────────────────
@app.post("/api/generate")
async def submit_image(req: Request):
    body = await req.json()
    job_id = uuid.uuid4().hex[:12]
    job = {
        "id": job_id, "status": "pending",
        "job_type":    "image",
        "prompt":      body.get("prompt", ""),
        "model":       body.get("model", ""),
        "size":        body.get("size", "1024x1024"),
        "api_mode":    body.get("api_mode", "openai"),
        "ollama_base": body.get("ollama_base", "http://localhost:11434"),
        "seed":        body.get("seed"),
        "steps":       body.get("steps"),
        "created":     time.time(),
        "client_id":   get_client_id(req),
    }
    write_job(job)
    task = asyncio.create_task(run_image_generation(job_id))
    _running_tasks[job_id] = task
    return {"job_id": job_id}


# ── Chat job submission ───────────────────────────────────────────────────────
# Accepts the full chat messages array and model config; runs the LLM
# server-side so the browser tab can be closed during generation.
# Chunks are streamed via SSE as "chat-chunk" events; completion as "chat-done".
# The job stores accumulated reply + updated messages so the frontend can
# restore history correctly on reconnect.
@app.post("/api/generate-chat")
async def submit_chat(req: Request):
    body = await req.json()
    job_id = uuid.uuid4().hex[:12]
    job = {
        "id":          job_id,
        "status":      "pending",
        "job_type":    "chat",
        # chat_id ties this job back to a specific frontend chat session so
        # the frontend can find "my pending job" on reconnect without
        # iterating all jobs.
        "chat_id":     body.get("chat_id", ""),
        "messages":    body.get("messages", []),
        "model":       body.get("model", ""),
        "api_flavor":  body.get("api_flavor", "ollama"),
        "ollama_base": body.get("ollama_base", "http://localhost:11434"),
        "options":     body.get("options", {}),
        "tools":        body.get("tools", []),
        "tools_config": body.get("tools_config", []),
        # Anti-runaway cap for sequential tool-call rounds, mirrored from the
        # client's Settings → Tool Calling (config.js MAX_TOOL_ROUNDS).
        # 0 = unlimited. Falls back to 5 if the client omits it.
        "max_tool_rounds": int(body.get("max_tool_rounds", 5)),
        # Image-gen settings for the built-in generate_image tool (see
        # _execute_generate_image_tool below) — mirrors Settings → Image on
        # the frontend so the model can generate images mid-chat without a
        # separate external tool config, correctly scoped to *this* client.
        "image_model":       body.get("image_model", ""),
        "image_api_mode":    body.get("image_api_mode", "openai"),
        "image_ollama_base": body.get("image_ollama_base", ""),
        "image_default_width":  int(body.get("image_default_width")  or 512),
        "image_default_height": int(body.get("image_default_height") or 512),
        "think":        bool(body.get("think", False)),
        "created":     time.time(),
        "client_id":   get_client_id(req),
        # Accumulated reply text so far (grown by the worker)
        "reply":       "",
        # The full messages array as updated by the worker (including
        # tool-call round-trips).  On completion, the frontend uses this
        # to replace its local chatHistory.
        "final_messages": [],
        # Tool call summaries for UI reconstruction on reconnect
        "tool_summaries": [],
        # RAG context injection — populated when frontend sends a collection name
        "rag_collection":   body.get("rag_collection", ""),
        "rag_qdrant_url":   body.get("rag_qdrant_url",  "http://localhost:6333"),
        "rag_embed_model":  body.get("rag_embed_model", "nomic-embed-text"),
        "rag_embed_flavor": body.get("rag_embed_flavor","ollama"),
        "rag_top_k":        int(body.get("rag_top_k", 5)),
        # RAG-as-a-tool — populated when the frontend's RAG mode is "Tool".
        # Unlike rag_collection above (auto-injected every turn), this
        # collection is only queried when the model explicitly calls the
        # rag_search tool, with a query/top_k the model itself decides.
        "rag_tool_collection":   body.get("rag_tool_collection", ""),
        "rag_tool_qdrant_url":   body.get("rag_tool_qdrant_url",  "http://localhost:6333"),
        "rag_tool_embed_model":  body.get("rag_tool_embed_model", "nomic-embed-text"),
        "rag_tool_embed_flavor": body.get("rag_tool_embed_flavor","ollama"),
        # Field whitelist/blacklist for rag_search results (Settings → RAG).
        # mode: 'all' (default, no filtering) | 'whitelist' | 'blacklist'.
        # "text" and "score" are always kept regardless of filter — see
        # _filter_rag_fields() below.
        "rag_tool_field_filter_mode": body.get("rag_tool_field_filter_mode", "all"),
        "rag_tool_field_whitelist":   body.get("rag_tool_field_whitelist", []),
        "rag_tool_field_blacklist":   body.get("rag_tool_field_blacklist", []),
        # Long-term memory — embedding/Qdrant config for the save_memory /
        # search_memory tools when this job is executed server-side. Sent by
        # memory.js's getMemoryToolBackendPayload() (mirrors the RAG-tool
        # fields above — same idea, one global collection instead of a
        # user-picked one). Falls back to the rag_tool_*/rag_* fields at
        # execution time if omitted (see _memory_embed_config in the tool
        # executors below), so this is optional from the client's side.
        "memory_embed_model":  body.get("memory_embed_model", ""),
        "memory_embed_flavor": body.get("memory_embed_flavor", ""),
        "memory_qdrant_url":   body.get("memory_qdrant_url", ""),
        # Minimum cosine similarity score a memory must clear to be
        # injected/returned — filters out low-relevance noise. 0 = no filter.
        "memory_min_score":    float(body.get("memory_min_score", 0.0) or 0.0),
    }
    write_job(job)
    task = asyncio.create_task(run_chat_generation(job_id))
    _running_tasks[job_id] = task
    return {"job_id": job_id}



@app.get("/api/jobs")
def list_jobs(req: Request):
    cid = get_client_id(req)
    jobs = []
    for p in JOBS_DIR.glob("*.json"):
        try:
            j = json.loads(p.read_text())
            if j.get("client_id") != cid: continue
            strip = {"b64", "result_text", "messages", "final_messages", "reply"}
            jobs.append({k: v for k, v in j.items() if k not in strip})
        except Exception: pass
    return {"jobs": sorted(jobs, key=lambda j: j.get("created", 0))}


@app.get("/api/jobs/{job_id}")
def get_job(job_id: str, req: Request):
    job = read_job(job_id)
    if not job or job.get("client_id") != get_client_id(req):
        return JSONResponse({"error": "not found"}, status_code=404)
    return job


@app.get("/api/jobs/{job_id}/files")
def list_job_files(job_id: str, req: Request):
    """List files saved from <jarvis_file> tags for a given job.
    Files are never deleted automatically — this endpoint just reports what exists."""
    job = read_job(job_id)
    if not job or job.get("client_id") != get_client_id(req):
        return JSONResponse({"error": "not found"}, status_code=404)
    job_files_dir = JARVIS_FILES_DIR / job_id
    if not job_files_dir.exists():
        return {"job_id": job_id, "files": []}
    files = [
        {"name": f.name, "path": str(f), "size": f.stat().st_size}
        for f in sorted(job_files_dir.iterdir())
        if f.is_file()
    ]
    return {"job_id": job_id, "files": files}


@app.delete("/api/jobs/{job_id}")
def delete_job(job_id: str, req: Request):
    job = read_job(job_id)
    if not job or job.get("client_id") != get_client_id(req):
        return JSONResponse({"error": "not found"}, status_code=404)
    job_path(job_id).unlink(missing_ok=True)
    _cleanup_reply_file(job_id)
    return {"deleted": job_id}


@app.post("/api/jobs/{job_id}/cancel")
async def cancel_job(job_id: str, req: Request):
    job = read_job(job_id)
    if not job or job.get("client_id") != get_client_id(req):
        return JSONResponse({"error": "not found"}, status_code=404)
    if job["status"] not in ("pending", "running"):
        return JSONResponse({"error": "not cancellable", "status": job["status"]}, status_code=400)
    task = _running_tasks.get(job_id)
    if task and not task.done():
        task.cancel()
    else:
        job["status"] = "cancelled"
        write_job(job)
        cid = job.get("client_id", "anonymous")
        await notify({"type": "cancelled", "job_id": job_id, "client_id": cid})
    return {"cancelled": job_id}


# ── SSE: global event stream ──────────────────────────────────────────────────
@app.get("/api/events")
async def sse_stream(req: Request, cid: str = Query(default="anonymous")):
    q = asyncio.Queue()
    _sse_queues.append((cid, q))

    async def stream():
        try:
            yield 'data: {"type":"connected"}\n\n'
            while True:
                if _shutdown_event.is_set():
                    break
                if await req.is_disconnected():
                    break
                try:
                    # Short timeout so we check disconnect/shutdown frequently.
                    # The ping keeps the connection alive through proxies.
                    event = await asyncio.wait_for(q.get(), timeout=5)
                    yield f"data: {json.dumps(event)}\n\n"
                except asyncio.TimeoutError:
                    yield ": ping\n\n"
        finally:
            entry = (cid, q)
            if entry in _sse_queues:
                _sse_queues.remove(entry)

    return StreamingResponse(
        stream(), media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )


# ── SSE: per-job replay stream ────────────────────────────────────────────────
# When the frontend reconnects after closing the tab mid-generation, it calls
# this endpoint to get all chunks emitted so far (buffered in job["reply"]) plus
# a live subscription to future chunks.  This lets the frontend reconstruct the
# streaming bubble without missing anything.
@app.get("/api/jobs/{job_id}/stream")
async def job_stream(job_id: str, req: Request):
    job = read_job(job_id)
    if not job:
        return JSONResponse({"error": "not found"}, status_code=404)

    cid = job.get("client_id", "anonymous")

    async def stream():
        # 1. Replay everything accumulated so far as catch-up chunks.
        # Read from the lightweight sidecar first (more current during streaming),
        # falling back to the full job fields if the sidecar is absent.
        fast_reply, fast_thinking = read_reply_fast(job_id)
        existing_thinking = fast_thinking or job.get("thinking", "")
        existing_reply    = fast_reply    or job.get("reply", "")
        if existing_thinking:
            yield f"data: {json.dumps({'type': 'chat-chunk', 'job_id': job_id, 'chunk': '', 'thinking': existing_thinking, 'replay': True})}\n\n"
        if existing_reply:
            yield f"data: {json.dumps({'type': 'chat-chunk', 'job_id': job_id, 'chunk': existing_reply, 'replay': True})}\n\n"

        # 2. If job is already finished, send the terminal event and close
        status = job.get("status", "")
        if status == "done":
            yield f"data: {json.dumps({'type': 'chat-done', 'job_id': job_id, 'gen_time': job.get('gen_time',''), 'replay': True})}\n\n"
            return
        if status in ("cancelled", "error"):
            yield f"data: {json.dumps({'type': 'chat-error', 'job_id': job_id, 'error': job.get('error', status), 'replay': True})}\n\n"
            return

        # 3. Job still running — subscribe to live events
        q = asyncio.Queue()
        _sse_queues.append((cid, q))
        try:
            while True:
                if _shutdown_event.is_set():
                    break
                if await req.is_disconnected():
                    break
                try:
                    event = await asyncio.wait_for(q.get(), timeout=5)
                    # Only forward events for this specific job
                    if event.get("job_id") == job_id:
                        yield f"data: {json.dumps(event)}\n\n"
                        if event.get("type") in ("chat-done", "chat-error", "cancelled", "error"):
                            break
                except asyncio.TimeoutError:
                    yield ": ping\n\n"
        finally:
            entry = (cid, q)
            if entry in _sse_queues:
                _sse_queues.remove(entry)

    return StreamingResponse(
        stream(), media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )


async def _emit_progress_loop(job_id: str, cid: str, start: float):
    """Emit elapsed-time progress events every 3 s while the job is running."""
    try:
        while True:
            await asyncio.sleep(3)
            job = read_job(job_id)
            if not job or job["status"] not in ("running", "pending"):
                break
            elapsed = round(time.time() - start, 1)
            await notify({"type": "progress", "job_id": job_id, "elapsed": elapsed, "client_id": cid})
    except asyncio.CancelledError:
        pass


# ── Image generation worker ───────────────────────────────────────────────────
async def run_image_generation(job_id: str):
    job = read_job(job_id)
    if not job: return
    cid = job.get("client_id", "anonymous")
    progress_task = None

    try:
        async with _ollama_sem:
            job = read_job(job_id)
            if not job or job.get("status") == "cancelled":
                return

            job["status"] = "running"
            job["started"] = time.time()
            write_job(job)
            await notify({"type": "status", "job_id": job_id, "status": "running", "client_id": cid})
            progress_task = asyncio.create_task(_emit_progress_loop(job_id, cid, job["started"]))

            try:
                ollama = job["ollama_base"].rstrip("/")
                b64 = None; gen_time = None
                
                # FIX: Set infinite read/pool timeout limits for huge matrix initializations
                timeout_config = httpx.Timeout(timeout=30.0, read=None, pool=None)
                
                async with httpx.AsyncClient(timeout=timeout_config) as client:
                    if job["api_mode"] == "native":
                        w, h = (int(x) for x in job["size"].split("x"))
                        body = {"model": job["model"], "prompt": job["prompt"], "width": w, "height": h}
                        if job.get("steps"):
                            body["steps"] = job["steps"]
                        if job.get("seed"):
                            body["options"] = {"seed": job["seed"]}
                        async with client.stream("POST", f"{ollama}/api/generate", json=body) as resp:
                            resp.raise_for_status()
                            async for line in resp.aiter_lines():
                                if not line.strip(): continue
                                try:
                                    o = json.loads(line)
                                    if isinstance(o.get("completed"), int) and isinstance(o.get("total"), int) and o["total"] > 0:
                                        elapsed = round(time.time() - job["started"], 1)
                                        await notify({"type": "progress", "job_id": job_id, "step": o["completed"], "steps_total": o["total"], "elapsed": elapsed, "client_id": cid})
                                    if o.get("done"):
                                        if o.get("total_duration"):
                                            gen_time = f"{o['total_duration']/1e9:.1f}s"
                                        if o.get("image"):
                                            b64 = o["image"]
                                except json.JSONDecodeError:
                                    # Skip raw log/heartbeat line data during cold generation spikes without crashing stream iterator
                                    pass
                    else:
                        resp = await client.post(f"{ollama}/v1/images/generations", json={"model": job["model"], "prompt": job["prompt"], "n": 1, "size": job["size"], "response_format": "b64_json"})
                        resp.raise_for_status()
                        items = resp.json().get("data", [])
                        if items:
                            b64 = items[0].get("b64_json", "")
                        if not b64:
                            w, h = (int(x) for x in job["size"].split("x"))
                            body = {"model": job["model"], "prompt": job["prompt"], "width": w, "height": h}
                            if job.get("steps"):
                                body["steps"] = job["steps"]
                            if job.get("seed"):
                                body["options"] = {"seed": job["seed"]}
                            async with client.stream("POST", f"{ollama}/api/generate", json=body) as resp:
                                resp.raise_for_status()
                                async for line in resp.aiter_lines():
                                    if not line.strip(): continue
                                    try:
                                        o = json.loads(line)
                                        if isinstance(o.get("completed"), int) and isinstance(o.get("total"), int) and o["total"] > 0:
                                            elapsed = round(time.time() - job["started"], 1)
                                            await notify({"type": "progress", "job_id": job_id, "step": o["completed"], "steps_total": o["total"], "elapsed": elapsed, "client_id": cid})
                                        if o.get("done"):
                                            if o.get("total_duration"):
                                                gen_time = f"{o['total_duration']/1e9:.1f}s"
                                            if o.get("image"):
                                                b64 = o["image"]
                                    except json.JSONDecodeError:
                                        # Skip raw log/heartbeat line data during cold generation spikes without crashing stream iterator
                                        pass
                            
                if not b64:
                    raise ValueError("No image returned by model")
                
                job["status"] = "done"; job["b64"] = b64
                job["gen_time"] = gen_time or f"{time.time()-job['started']:.1f}s"
                job["finished"] = time.time(); write_job(job)
                await notify({"type": "done", "job_id": job_id, "client_id": cid})

            except Exception as e:
                import traceback
                traceback.print_exc()
                job = read_job(job_id) or job
                if job.get("status") not in ("done", "cancelled"):
                    job["status"] = "error"; job["error"] = str(e); write_job(job)
                    await notify({"type": "error", "job_id": job_id, "error": str(e), "client_id": cid})
            finally:
                if progress_task:
                    progress_task.cancel()

    except asyncio.CancelledError:
        job = read_job(job_id) or job
        if job.get("status") not in ("done", "error", "cancelled"):
            job["status"] = "cancelled"; write_job(job)
            await notify({"type": "cancelled", "job_id": job_id, "client_id": cid})

    finally:
        _running_tasks.pop(job_id, None)


# ── Chat generation worker ────────────────────────────────────────────────────
# Full backend chat: handles streaming, tool-call round-trips, and persists
# accumulated reply + updated messages array so reconnecting tabs can catch up.
async def run_chat_generation(job_id: str):
    job = read_job(job_id)
    if not job: return
    cid      = job.get("client_id", "anonymous")
    flavor   = job.get("api_flavor", "ollama")
    base     = job.get("ollama_base", "http://localhost:11434").rstrip("/")
    model    = job.get("model", "")
    messages = list(job.get("messages", []))
    options  = job.get("options", {})
    tools    = job.get("tools", [])
    think    = bool(job.get("think", False))
    progress_task = None

    # ── RAG context injection ─────────────────────────────────────────────────
    # If the job specifies a rag_collection, embed the last user message,
    # retrieve top-k chunks from Qdrant, and prepend them to the system prompt.
    # The retrieved snippets are also sent back as a "rag_chunks" event so the
    # frontend can render a collapsible context card (like tool calls).
    _rag_chunks: list[dict] = []
    _rag_collection = job.get("rag_collection", "")
    if _rag_collection:
        # Find the last user message to use as the search query
        _last_user = next(
            (m["content"] for m in reversed(messages) if m.get("role") == "user"),
            None,
        )
        if _last_user and isinstance(_last_user, str):
            _rag_chunks = await rag_search_chunks(
                query        = _last_user[:2000],   # cap query length
                collection   = _rag_collection,
                embed_model  = job.get("rag_embed_model",  "nomic-embed-text"),
                embed_flavor = job.get("rag_embed_flavor", "ollama"),
                ollama_base  = base,
                qdrant_url   = job.get("rag_qdrant_url",  "http://localhost:6333"),
                top_k        = job.get("rag_top_k", 5),
            )
        if _rag_chunks:
            # Build a context block to inject into the system prompt
            _ctx_block = (
                "\n\n---\n"
                "## Relevant context retrieved from the knowledge base\n"
                "Use the following excerpts to answer the user's question. "
                "Cite the source filename when referencing them.\n\n"
            )
            for i, c in enumerate(_rag_chunks, 1):
                _src  = c.get("source", "") or ""
                _cidx = c.get("chunk_index")
                _meta = f" (chunk {_cidx})" if _cidx is not None else ""
                _ctx_block += (
                    f"### [{i}]{(' ' + _src) if _src else ''}{_meta} "
                    f"score {c.get('score', 0):.3f}\n{c.get('text', '')}\n\n"
                )
            _ctx_block += "---"

            # Inject into the system message (first message if role=system, else prepend one)
            if messages and messages[0].get("role") == "system":
                messages[0] = dict(messages[0])   # don't mutate original
                messages[0]["content"] = (messages[0]["content"] or "") + _ctx_block
            else:
                messages.insert(0, {"role": "system", "content": _ctx_block.strip()})

        # Emit a rag-chunks SSE event so the frontend can render a context card
        # (displayed similarly to tool-call summaries, before the reply)
        await notify({
            "type":       "rag-chunks",
            "job_id":     job_id,
            "chat_id":    job.get("chat_id", ""),
            "client_id":  cid,
            "collection": _rag_collection,
            "chunks":     _rag_chunks,   # [] when no collection or no hits
        })

    try:
        job = read_job(job_id)
        if not job or job.get("status") == "cancelled":
            return

        job["status"]  = "running"
        job["started"] = time.time()
        write_job(job)
        await notify({
            "type": "status", "job_id": job_id, "status": "running",
            "job_type": "chat", "chat_id": job.get("chat_id", ""),
            "client_id": cid,
        })
        progress_task = asyncio.create_task(_emit_progress_loop(job_id, cid, job["started"]))

        try:
            reply = ""
            thinking_reply = ""   # accumulated thinking text (separate from reply)
            all_tool_summaries = []  # [{name, args, result, error}]

            # 0 = unlimited (job["max_tool_rounds"] set from client payload above)
            MAX_TOOL_ROUNDS = int(job.get("max_tool_rounds", 5))
            tool_round = 0

            # FIX: Standardized chat timeouts to survive long reasoning chains
            chat_timeout = httpx.Timeout(timeout=30.0, read=None, pool=None)

            async with httpx.AsyncClient(timeout=chat_timeout) as client:
                while True:
                    # ── Build request body ────────────────────────────────
                    if flavor == "ollama":
                        body: dict = {
                            "model":    model,
                            "messages": messages,
                            "stream":   True,
                            "options":  options,
                            "think":    think,
                        }
                        if tools:
                            body["tools"] = tools
                        stream_url = f"{base}/api/chat"
                    else:
                        # OpenAI-compatible
                        body = {
                            "model":    model,
                            "messages": messages,
                            "stream":   True,
                        }
                        if "temperature"    in options: body["temperature"]      = options["temperature"]
                        if "top_p"          in options: body["top_p"]            = options["top_p"]
                        if "num_ctx"        in options: body["max_tokens"]       = options["num_ctx"]
                        if "repeat_penalty" in options: body["frequency_penalty"] = options["repeat_penalty"] - 1.0
                        if "top_k"          in options: body["top_k"]            = options["top_k"]
                        if tools:
                            body["tools"] = tools
                        stream_url = f"{base}/v1/chat/completions"

                    # ── Stream one LLM turn ───────────────────────────────
                    turn_reply   = ""
                    oai_tool_acc = {}   # OpenAI streaming tool-call deltas
                    ollama_tools = None # Ollama final tool_calls
                    last_raw_chunk = None  # Ollama done-chunk for stats

                    async with client.stream("POST", stream_url, json=body) as resp:
                        resp.raise_for_status()
                        async for line in resp.aiter_lines():
                            if not line.strip():
                                continue
                            if line.strip() == "data: [DONE]":
                                continue
                            raw = line[6:] if line.startswith("data: ") else line
                            try:
                                o = json.loads(raw)
                            except Exception:
                                continue

                            chunk = ""
                            thinking_chunk = ""
                            if flavor == "ollama":
                                msg = o.get("message") or {}
                                chunk = msg.get("content") or ""
                                # Newer Ollama versions (0.5+) expose thinking text
                                # in a dedicated message.thinking field, separate
                                # from message.content which arrives pre-stripped.
                                thinking_chunk = msg.get("thinking") or ""
                                # Capture the done chunk for stats (total_duration etc.)
                                if o.get("done"):
                                    last_raw_chunk = o
                                # Capture tool_calls from any chunk — when thinking is
                                # enabled Ollama emits tool_calls on a non-done chunk.
                                if msg.get("tool_calls"):
                                    ollama_tools = msg["tool_calls"]
                            else:
                                delta = ((o.get("choices") or [{}])[0].get("delta") or {})
                                chunk = delta.get("content") or ""
                                # DeepSeek uses reasoning_content; some providers use thinking
                                thinking_chunk = delta.get("thinking") or delta.get("reasoning_content") or ""
                                # Accumulate OpenAI streaming tool-call deltas
                                if delta.get("tool_calls"):
                                    for tc in delta["tool_calls"]:
                                        idx = tc.get("index", 0)
                                        if idx not in oai_tool_acc:
                                            oai_tool_acc[idx] = {"id": "", "type": "function",
                                                                  "function": {"name": "", "arguments": ""}}
                                        if tc.get("id"):
                                            oai_tool_acc[idx]["id"] = tc["id"]
                                        if tc.get("function", {}).get("name"):
                                            oai_tool_acc[idx]["function"]["name"] += tc["function"]["name"]
                                        if tc.get("function", {}).get("arguments"):
                                            oai_tool_acc[idx]["function"]["arguments"] += tc["function"]["arguments"]

                            if thinking_chunk:
                                thinking_reply += thinking_chunk
                                # Persist thinking cheaply (sidecar file, no image data)
                                write_reply_fast(job_id, reply, thinking_reply)
                                await notify({
                                    "type":    "chat-chunk",
                                    "job_id":  job_id,
                                    "chat_id": job.get("chat_id", ""),
                                    "chunk":   "",
                                    "thinking": thinking_chunk,
                                    "client_id": cid,
                                })

                            if chunk:
                                turn_reply += chunk
                                reply      += chunk
                                # Persist accumulated reply cheaply (sidecar file, no image data).
                                # The full job JSON (which may contain MB of base64 images) is
                                # only rewritten on state transitions, not on every token.
                                write_reply_fast(job_id, reply, thinking_reply)
                                await notify({
                                    "type":    "chat-chunk",
                                    "job_id":  job_id,
                                    "chat_id": job.get("chat_id", ""),
                                    "chunk":   chunk,
                                    "client_id": cid,
                                })

                    # ── Resolve tool calls ────────────────────────────────
                    tool_calls = []
                    if oai_tool_acc:
                        tool_calls = [
                            {
                                "id": v.get("id") or f"call_{i}",
                                "function": {
                                    "name": v["function"]["name"],
                                    "arguments": _safe_parse_args(v["function"]["arguments"]),
                                }
                            }
                            for i, v in oai_tool_acc.items()
                            if v["function"].get("name")
                        ]
                    elif ollama_tools:
                        tool_calls = [
                            {
                                "id": f"call_{i}",
                                "function": {
                                    "name":      tc["function"]["name"],
                                    "arguments": tc["function"].get("arguments") or {},
                                }
                            }
                            for i, tc in enumerate(ollama_tools)
                        ]

                    if tool_calls and (MAX_TOOL_ROUNDS == 0 or tool_round < MAX_TOOL_ROUNDS) and tools:
                        # Let the frontend know tool execution is starting —
                        # in particular so it can show a loading placeholder
                        # for generate_image calls while they're in flight.
                        # The full results follow later via chat-tool-calls;
                        # this is fire-and-forget, best-effort UI only.
                        await notify({
                            "type":      "chat-tool-start",
                            "job_id":    job_id,
                            "chat_id":   job.get("chat_id", ""),
                            "calls":     [{"name": tc["function"]["name"]} for tc in tool_calls],
                            "client_id": cid,
                        })

                        # Append assistant's tool-call message to history
                        if flavor == "ollama":
                            messages.append({
                                "role": "assistant", "content": "",
                                "tool_calls": [
                                    {"function": {"name": tc["function"]["name"],
                                                  "arguments": tc["function"]["arguments"]}}
                                    for tc in tool_calls
                                ]
                            })
                        else:
                            messages.append({
                                "role": "assistant", "content": None,
                                "tool_calls": [
                                    {"id": tc["id"], "type": "function",
                                     "function": {"name": tc["function"]["name"],
                                                  "arguments": json.dumps(tc["function"]["arguments"])}}
                                    for tc in tool_calls
                                ]
                            })

                        # Execute each tool. rag_search runs server-side against
                        # Qdrant; HTTP tools run via their saved config. MCP tools
                        # can't run server-side yet.
                        round_images = []  # THIS round's generate_image results only —
                                            # see note on the notify() call below for why
                                            # this must stay separate from all_tool_summaries.
                        for tc in tool_calls:
                            result_str, error = await _execute_tool_call(tc, job, job_id, cid)
                            # generate_image returns a `[[JARVIS_IMAGE:id]]` marker
                            # (see _execute_generate_image_tool) so the frontend can
                            # detect and render the image inline. Strip it before
                            # committing to the model-facing message log, so it
                            # doesn't linger in every future turn's conversation
                            # history sent back to the model.
                            img_match = _re.match(r"\[\[JARVIS_IMAGE:([A-Za-z0-9_]+)\]\]", result_str)
                            model_facing_result = _re.sub(r"\[\[JARVIS_IMAGE:[A-Za-z0-9_]+\]\]\s*", "", result_str).strip() or result_str
                            if flavor == "ollama":
                                messages.append({"role": "tool", "content": model_facing_result})
                            else:
                                messages.append({"role": "tool", "tool_call_id": tc["id"], "content": model_facing_result})

                            all_tool_summaries.append({
                                "name":   tc["function"]["name"],
                                "args":   tc["function"]["arguments"],
                                "result": result_str,
                                "error":  error,
                            })

                            # The sub-job created by _execute_generate_image_tool is
                            # fully consumed here — its b64 travels directly in this
                            # SSE event below, so the frontend never needs to fetch
                            # (and race/poll for) it separately. Delete it once read.
                            if img_match and not error:
                                sub_job_id = img_match.group(1)
                                sub_job = read_job(sub_job_id)
                                if sub_job and sub_job.get("b64"):
                                    round_images.append({
                                        "id":       sub_job_id,
                                        "b64":      sub_job["b64"],
                                        "prompt":   sub_job.get("prompt", ""),
                                        "model":    sub_job.get("model", ""),
                                        "size":     sub_job.get("size", ""),
                                        "gen_time": sub_job.get("gen_time", ""),
                                    })
                                job_path(sub_job_id).unlink(missing_ok=True)

                        # Notify frontend about tool call activity. tool_summaries is
                        # intentionally the FULL cumulative list across all rounds so
                        # far — _appendToolSummary() on the frontend redraws its whole
                        # widget from scratch on every call and expects the complete
                        # history. images, by contrast, is scoped to THIS round only:
                        # sending the cumulative image list here would make the
                        # frontend re-process (and re-render/duplicate) earlier
                        # rounds' already-displayed images on every subsequent round.
                        await notify({
                            "type":           "chat-tool-calls",
                            "job_id":         job_id,
                            "chat_id":        job.get("chat_id", ""),
                            "tool_summaries": all_tool_summaries,
                            "images":         round_images,
                            "client_id":      cid,
                        })

                        tool_round += 1
                        continue  # loop for model's continuation after tools

                    # No tool calls (or cap reached) — done with this generation
                    break

            # Append final assistant message to messages
            messages.append({"role": "assistant", "content": reply})

            gen_time = f"{time.time() - job['started']:.1f}s"
            job = read_job(job_id) or job
            job["status"]          = "done"
            job["reply"]           = reply
            job["thinking"]        = thinking_reply
            job["final_messages"]  = messages
            job["tool_summaries"]  = all_tool_summaries
            job["gen_time"]        = gen_time
            job["raw_stats"]       = last_raw_chunk  # Ollama stats for frontend renderStatsBar
            job["finished"]        = time.time()
            # Extract and save any <jarvis_file> tags from the final reply
            saved_files = save_jarvis_files(job_id, reply)
            if saved_files:
                job["jarvis_files"] = saved_files
            write_job(job)
            _cleanup_reply_file(job_id)   # sidecar merged into full job; no longer needed
            await notify({
                "type":           "chat-done",
                "job_id":         job_id,
                "chat_id":        job.get("chat_id", ""),
                "gen_time":       gen_time,
                "raw_stats":      last_raw_chunk,
                "tool_summaries": all_tool_summaries,
                "client_id":      cid,
            })

        except asyncio.CancelledError:
            job = read_job(job_id) or job
            job["status"] = "cancelled"
            # Save any jarvis_file tags present in the partial reply
            partial_reply = job.get("reply", "")
            saved_files = save_jarvis_files(job_id, partial_reply)
            if saved_files:
                job["jarvis_files"] = saved_files
            write_job(job)
            _cleanup_reply_file(job_id)
            await notify({"type": "cancelled", "job_id": job_id,
                          "chat_id": job.get("chat_id", ""), "client_id": cid})

        except Exception as e:
            job = read_job(job_id) or job
            job["status"] = "error"
            job["error"]  = str(e)
            # Save any jarvis_file tags present in the partial reply
            partial_reply = job.get("reply", "")
            saved_files = save_jarvis_files(job_id, partial_reply)
            if saved_files:
                job["jarvis_files"] = saved_files
            write_job(job)
            _cleanup_reply_file(job_id)
            await notify({"type": "chat-error", "job_id": job_id,
                          "chat_id": job.get("chat_id", ""),
                          "error": str(e), "job_type": "chat", "client_id": cid})

        finally:
            if progress_task:
                progress_task.cancel()

    except asyncio.CancelledError:
        job = read_job(job_id) or job
        if job.get("status") not in ("done", "error", "cancelled"):
            job["status"] = "cancelled"
            write_job(job)
            await notify({"type": "cancelled", "job_id": job_id, "client_id": cid})

    finally:
        _running_tasks.pop(job_id, None)


def _safe_parse_args(arguments_str):
    try:
        return json.loads(arguments_str or "{}")
    except Exception:
        return {}


async def _execute_tool_call(tc: dict, job: dict, job_id: str, cid: str) -> tuple[str, bool]:
    """Dispatch a single tool call to the right executor:
    rag_search (built-in, server-side Qdrant retrieval), generate_image
    (built-in, native image generation), or an HTTP tool from the job's
    saved tools_config. MCP tools are rejected by _execute_http_tool with a
    clear error message."""
    name = tc["function"]["name"]
    if name == "rag_search":
        return await _execute_rag_search_tool(tc, job, job_id, cid)
    if name == "generate_image":
        return await _execute_generate_image_tool(tc, job, job_id, cid)
    if name == "save_memory":
        return await _execute_save_memory_tool(tc, job, job_id, cid)
    if name == "search_memory":
        return await _execute_search_memory_tool(tc, job, job_id, cid)
    if name == "update_memory":
        return await _execute_update_memory_tool(tc, job, job_id, cid)
    return await _execute_http_tool(tc, job)


async def _execute_generate_image_tool(tc: dict, job: dict, job_id: str, cid: str) -> tuple[str, bool]:
    """Executes the built-in generate_image tool natively, server-side, by
    reusing run_image_generation() directly instead of routing through an
    external HTTP tool. This fixes two things an external-tool config can't:
      1. The resulting job is scoped to *cid* — the actual chat client for
         this request — instead of whatever client-id a manually-configured
         HTTP tool's headers happen to hardcode, which previously meant the
         job was invisible to the browser that asked for it (/api/jobs and
         /api/jobs/{id} both filter by client_id).
      2. job_type is "tool_image", not "image", so the browser's generic
         gallery-sync (syncBackendJobs(), which only picks up job_type ==
         "image") never auto-adds it — saving to the gallery stays opt-in,
         via the button on the inline chat card.
    Returns a `[[JARVIS_IMAGE:job_id]]` marker (mirroring the local/direct
    tool-call path in chat.js) instead of raw base64, so the (possibly
    multi-MB) image never rides along in the tool-result text sent back to
    the model on this or future turns.
    """
    args = tc["function"]["arguments"] or {}

    prompt = str(args.get("prompt", "") or "").strip()
    default_w = int(job.get("image_default_width")  or 512)
    default_h = int(job.get("image_default_height") or 512)
    raw_w = args.get("width")
    raw_h = args.get("height")
    try:
        width  = default_w if raw_w in (None, "") else int(raw_w)
        height = default_h if raw_h in (None, "") else int(raw_h)
    except (TypeError, ValueError):
        return "Error: width and height must be integers.", True

    if not prompt:
        return "Error: prompt must not be empty.", True
    if width % 16 != 0 or height % 16 != 0:
        return "Error: width and height must both be exact multiples of 16.", True
    if width > 1920 or height > 1920:
        return "Error: width and height must not exceed 1920.", True
    if width < 64 or height < 64:
        return "Error: width and height must be at least 64.", True

    image_model = job.get("image_model") or ""
    if not image_model:
        return "Error: no image model configured (check Settings → Image).", True

    sub_job_id = uuid.uuid4().hex[:12]
    sub_job = {
        "id": sub_job_id, "status": "pending",
        "job_type":    "tool_image",   # excluded from gallery auto-sync — see docstring
        "prompt":      prompt,
        "model":       image_model,
        "size":        f"{width}x{height}",
        "api_mode":    job.get("image_api_mode", "openai"),
        "ollama_base": job.get("image_ollama_base") or job.get("ollama_base", "http://localhost:11434"),
        "seed": None, "steps": None,
        "created":   time.time(),
        "client_id": cid,
    }
    write_job(sub_job)

    # Announce the sub-job id immediately — run_image_generation() below emits
    # its 'progress'/'status' SSE events keyed by sub_job_id, not job_id (this
    # tool call's parent chat job). Without this, the frontend has no way to
    # know sub_job_id exists until the image is already finished (it only
    # otherwise appears in the later chat-tool-calls event), so any progress
    # events emitted during generation land on an id nobody's listening for.
    await notify({
        "type":       "chat-tool-image-job",
        "job_id":     job_id,
        "chat_id":    job.get("chat_id", ""),
        "sub_job_id": sub_job_id,
        "client_id":  cid,
    })

    try:
        await run_image_generation(sub_job_id)
    except Exception as e:
        return f"Error generating image: {e}", True

    result_job = read_job(sub_job_id)
    if not result_job or result_job.get("status") != "done" or not result_job.get("b64"):
        err = (result_job or {}).get("error", "Image generation failed.")
        return f"Error generating image: {err}", True

    return (
        f"[[JARVIS_IMAGE:{sub_job_id}]] Image generated successfully and displayed to the user in the chat.",
        False,
    )


async def _execute_rag_search_tool(tc: dict, job: dict, job_id: str, cid: str) -> tuple[str, bool]:
    """Executes the model-driven rag_search tool call: embeds the model's own
    query, retrieves top_k chunks from the collection configured for this job
    (rag_tool_collection), and notifies the frontend so it can render a
    context card — same as the always-inject path, but per tool call.
    Runs entirely server-side since backend-offloaded jobs may keep running
    after the tab is closed."""
    args       = tc["function"]["arguments"] or {}
    collection = job.get("rag_tool_collection", "")
    if not collection:
        return json.dumps({"error": "No RAG collection is configured for this tool."}), True

    query = str(args.get("query", "") or "")[:2000]
    if not query:
        return json.dumps({"error": "The 'query' parameter is required."}), True

    try:
        top_k = int(args.get("top_k", 5))
    except (TypeError, ValueError):
        top_k = 5
    top_k = max(1, min(20, top_k))

    try:
        chunks = await rag_search_chunks(
            query        = query,
            collection   = collection,
            embed_model  = job.get("rag_tool_embed_model",  "nomic-embed-text"),
            embed_flavor = job.get("rag_tool_embed_flavor", "ollama"),
            ollama_base  = job.get("ollama_base", "http://localhost:11434"),
            qdrant_url   = job.get("rag_tool_qdrant_url",  "http://localhost:6333"),
            top_k        = top_k,
        )
    except Exception as e:
        return json.dumps({"error": f"RAG search failed: {e}"}), True

    # Notify the frontend so it can render a context card for this call,
    # the same way always-mode retrieval and HTTP tool calls are surfaced.
    await notify({
        "type":       "rag-chunks",
        "job_id":     job_id,
        "chat_id":    job.get("chat_id", ""),
        "client_id":  cid,
        "collection": collection,
        "chunks":     chunks,
    })

    if not chunks:
        return json.dumps({
            "query": query, "top_k": top_k, "results": [],
            "note": "No relevant chunks found for this query.",
        }), False

    # Pass through the full chunk payload by default (source, score, text,
    # plus whatever extra metadata fields the collection stores — ril_number,
    # page, section_title, module_title, breadcrumb, etc.), then apply the
    # user's whitelist/blacklist (Settings → RAG) if one is configured.
    filtered = [_filter_rag_fields(c, job) for c in chunks]
    return json.dumps({"query": query, "top_k": top_k, "results": filtered}), False


def _filter_rag_fields(chunk: dict, job: dict) -> dict:
    """Applies the user-configured field whitelist/blacklist (Settings → RAG)
    to a single resolved chunk dict before it's sent to the model. "text" and
    "score" are always kept — dropping the chunk's actual content or its
    relevance score would defeat the tool's purpose, so the filter only
    applies to the rest of the metadata fields."""
    mode = job.get("rag_tool_field_filter_mode", "all")
    if mode not in ("whitelist", "blacklist"):
        return chunk
    whitelist = set(job.get("rag_tool_field_whitelist") or [])
    blacklist = set(job.get("rag_tool_field_blacklist") or [])
    out = {}
    for k, v in chunk.items():
        if k in ("text", "score"):
            out[k] = v
            continue
        if mode == "whitelist" and k not in whitelist:
            continue
        if mode == "blacklist" and k in blacklist:
            continue
        out[k] = v
    return out


def _memory_embed_config(job: dict) -> dict:
    """Best-effort embedding/Qdrant config for memory tool calls. Prefers the
    dedicated memory_* fields (sent by memory.js's getMemoryToolBackendPayload,
    which mirrors the browser's actual RAG_EMBED_MODEL/RAG_QDRANT_URL settings —
    memory reuses the same embedding pipeline as RAG, see product notes),
    falling back to whatever RAG config the job happens to carry, then to
    hardcoded defaults so this never hard-fails for lack of config."""
    return {
        "embed_model":  job.get("memory_embed_model")  or job.get("rag_tool_embed_model")  or job.get("rag_embed_model")  or "nomic-embed-text",
        "embed_flavor": job.get("memory_embed_flavor") or job.get("rag_tool_embed_flavor") or job.get("rag_embed_flavor") or "ollama",
        "qdrant_url":   job.get("memory_qdrant_url")   or job.get("rag_tool_qdrant_url")   or job.get("rag_qdrant_url")   or "http://localhost:6333",
        "ollama_base":  job.get("ollama_base", "http://localhost:11434"),
    }


async def _execute_save_memory_tool(tc: dict, job: dict, job_id: str, cid: str) -> tuple[str, bool]:
    """Executes the model-driven save_memory tool call. Writes directly via
    memory_save() (imported from memory.py) instead of an HTTP round-trip to
    this same server — same reasoning as _execute_generate_image_tool reusing
    run_image_generation() directly. Notifies the frontend with a lightweight
    'memory-saved' event so a toast can confirm the write without needing a
    full context card (there are no retrieved chunks to show, just one save)."""
    args = tc["function"]["arguments"] or {}
    text = str(args.get("text", "") or "").strip()
    if not text:
        return json.dumps({"error": "The 'text' parameter is required and must not be empty."}), True

    tags = args.get("tags") or []
    if not isinstance(tags, list):
        tags = [str(tags)]
    confidence = args.get("confidence") or "stated"
    if confidence not in ("stated", "inferred"):
        confidence = "stated"

    cfg = _memory_embed_config(job)
    try:
        result = await memory_save(cid, SaveMemoryRequest(
            text=text[:4000],   # keep individual memories short and self-contained
            tags=[str(t) for t in tags][:10],
            source=job.get("chat_id", "") or "chat",
            confidence=confidence,
            embed_model=cfg["embed_model"], embed_flavor=cfg["embed_flavor"],
            ollama_base=cfg["ollama_base"], qdrant_url=cfg["qdrant_url"],
        ))
    except Exception as e:
        return json.dumps({"error": f"Failed to save memory: {e}"}), True

    await notify({
        "type":      "memory-saved",
        "job_id":    job_id,
        "chat_id":   job.get("chat_id", ""),
        "client_id": cid,
        "text":      text[:200],
        "indexed":   bool(result.get("indexed")),
    })

    return json.dumps({"saved": True, "id": result.get("id"), "indexed": result.get("indexed", False)}), False


async def _execute_search_memory_tool(tc: dict, job: dict, job_id: str, cid: str) -> tuple[str, bool]:
    """Executes the model-driven search_memory tool call — an explicit,
    on-demand lookup in addition to the automatic per-turn retrieval the
    frontend already performs (memory.js's injectMemoryContext). Useful when
    the model wants a different or more specific query than the user's
    latest message."""
    args  = tc["function"]["arguments"] or {}
    query = str(args.get("query", "") or "")[:2000]
    if not query:
        return json.dumps({"error": "The 'query' parameter is required."}), True
    try:
        top_k = int(args.get("top_k", 5))
    except (TypeError, ValueError):
        top_k = 5
    top_k = max(1, min(20, top_k))

    cfg = _memory_embed_config(job)
    try:
        results = await memory_search_chunks(
            client_id=cid, query=query, embed_model=cfg["embed_model"], embed_flavor=cfg["embed_flavor"],
            ollama_base=cfg["ollama_base"], qdrant_url=cfg["qdrant_url"], top_k=top_k,
            score_threshold=job.get("memory_min_score", 0.0),
        )
    except Exception as e:
        return json.dumps({"error": f"Memory search failed: {e}"}), True

    if not results:
        return json.dumps({"query": query, "top_k": top_k, "results": [], "note": "No matching memories found."}), False
    return json.dumps({"query": query, "top_k": top_k, "results": results}), False


async def _execute_update_memory_tool(tc: dict, job: dict, job_id: str, cid: str) -> tuple[str, bool]:
    """Executes the model-driven update_memory tool call. Requires an
    explicit entry `id` — by design there is no fuzzy/best-guess lookup
    here, so the model cannot accidentally overwrite the wrong memory.
    The intended flow (spelled out in the tool description, config.js) is:
    call search_memory first, read the id of the matching entry from its
    results, then call update_memory with that exact id."""
    args     = tc["function"]["arguments"] or {}
    entry_id = str(args.get("id", "") or "").strip()
    if not entry_id:
        return json.dumps({
            "error": "The 'id' parameter is required. Call search_memory first to find the "
                     "exact id of the memory you want to update — never guess an id."
        }), True

    text = args.get("text")
    tags = args.get("tags")
    if tags is not None and not isinstance(tags, list):
        tags = [str(tags)]
    confidence = args.get("confidence")
    if confidence is not None and confidence not in ("stated", "inferred"):
        confidence = None

    cfg = _memory_embed_config(job)
    try:
        result = await memory_update(cid, entry_id, UpdateMemoryRequest(
            text=(text.strip()[:4000] if isinstance(text, str) and text.strip() else None),
            tags=([str(t) for t in tags][:10] if tags is not None else None),
            confidence=confidence,
            embed_model=cfg["embed_model"], embed_flavor=cfg["embed_flavor"],
            ollama_base=cfg["ollama_base"], qdrant_url=cfg["qdrant_url"],
        ))
    except HTTPException as e:
        return json.dumps({"error": e.detail}), True
    except Exception as e:
        return json.dumps({"error": f"Failed to update memory: {e}"}), True

    await notify({
        "type":      "memory-saved",
        "job_id":    job_id,
        "chat_id":   job.get("chat_id", ""),
        "client_id": cid,
        "text":      (result.get("text") or "")[:200],
        "indexed":   bool(result.get("indexed")),
        "updated":   True,
    })

    return json.dumps({"updated": True, "id": entry_id, "indexed": result.get("indexed", False)}), False


async def _execute_http_tool(tc: dict, job: dict) -> tuple[str, bool]:
    """Execute a single tool call server-side using the tool config from the job.
    Returns (result_string, is_error).  MCP tools are not yet supported
    server-side; they log a clear error instead of crashing the job."""
    tools_config = job.get("tools_config", [])
    name = tc["function"]["name"]
    args = tc["function"]["arguments"] or {}

    # MCP tools (namespaced mcp__serverid__toolname) can't run server-side
    # because MCP servers typically bind to localhost and aren't reachable
    # from the backend container.  Return a clear message so the model knows.
    if name.startswith("mcp__"):
        return f"[Tool '{name}' is an MCP tool and cannot be executed server-side]", True

    # Find the matching HTTP tool config
    tool_conf = next((t for t in tools_config if t.get("type") == "http" and t.get("name") == name), None)
    if not tool_conf:
        return f"[Tool '{name}' not found in server-side config]", True

    try:
        method  = (tool_conf.get("method") or "GET").upper()
        url     = tool_conf.get("url", "")
        headers = dict(tool_conf.get("headers") or {})

        for k, v in args.items():
            url = url.replace(f"{{{{{k}}}}}", str(v))

        req_opts: dict = {"method": method, "url": url, "headers": headers, "timeout": 30}

        if method not in ("GET", "HEAD"):
            body_tpl = (tool_conf.get("bodyTemplate") or "").strip()
            if body_tpl:
                for k, v in args.items():
                    body_tpl = body_tpl.replace(f"{{{{{k}}}}}", str(v))
                req_opts["content"] = body_tpl
            else:
                req_opts["json"] = args
            if "Content-Type" not in headers:
                headers["Content-Type"] = "application/json"

        async with httpx.AsyncClient() as client:
            resp = await client.request(**req_opts)
            text = resp.text
            if not resp.is_success:
                return f"HTTP {resp.status_code}: {text[:300]}", True

        # Apply response filter if configured
        resp_filter = tool_conf.get("responseFilter")
        if resp_filter:
            text = _apply_response_filter(text, resp_filter)

        return text, False
    except Exception as e:
        return f"Tool error: {e}", True


def _apply_response_filter(text: str, filt: dict) -> str:
    pick  = [f.strip() for f in (filt.get("pick")  or "").split(",") if f.strip()]
    drop  = [f.strip() for f in (filt.get("drop")  or "").split(",") if f.strip()]
    limit = int(filt.get("limit") or 0)
    if not pick and not drop and not limit:
        return text
    try:
        data = json.loads(text)
    except Exception:
        return text
    target = data
    if filt.get("path"):
        # Split on unescaped dots (matching the frontend's regex behaviour)
        import re
        keys = re.split(r'(?<!\\)\.', filt["path"])
        for key in keys:
            key = key.replace(r'\.', '.')
            if target is None or not isinstance(target, dict):
                return text
            target = target.get(key)

    def filter_obj(obj):
        if not isinstance(obj, dict):
            return obj
        if pick:
            return {f: obj[f] for f in pick if f in obj}
        return {k: v for k, v in obj.items() if k not in drop}

    filtered = [filter_obj(i) for i in target] if isinstance(target, list) else filter_obj(target)
    if limit > 0 and isinstance(filtered, list):
        filtered = filtered[:limit]
    return json.dumps(filtered, indent=2)
