# ══ RAG.PY ════════════════════════════════════════════════════
# RAG (Retrieval-Augmented Generation) backend for Jarvis.
# Provides endpoints for:
#   - Collection management (create, list, delete)
#   - File ingestion (chunking + embedding via Ollama + upsert to Qdrant)
#   - Semantic search (query a collection, returns top-k chunks)
#
# Mount into server.py with:
#   from rag import rag_router
#   app.include_router(rag_router)
#
# Dependencies (add to requirements.txt / pip install):
#   qdrant-client>=1.9.0
#   httpx  (already present in server.py)
#   markitdown (already present)
#   pypdf (already present)

import asyncio, json, math, re, tempfile, os, time, uuid
from pathlib import Path
from typing import Any

import httpx
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Request
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel

# ── Optional heavy imports ────────────────────────────────────
try:
    from qdrant_client import QdrantClient, AsyncQdrantClient
    from qdrant_client import models as qmodels
    from qdrant_client.models import (
        Distance, VectorParams, PointStruct,
        Filter, FieldCondition, MatchValue,
    )
    _QDRANT_OK = True
except ImportError:
    _QDRANT_OK = False

try:
    from markitdown import MarkItDown
    _markitdown = MarkItDown()
except ImportError:
    _markitdown = None

# ── Payload field resolver ────────────────────────────────────────────────────
# Collections ingested by external tools (LangChain, LlamaIndex, custom scripts)
# use different field names than what this app writes. This function resolves
# the actual text and metadata from whatever fields are present in the payload.

# Candidate field names for the main text content, in priority order
_TEXT_FIELDS   = ["text", "content", "page_content", "body", "chunk", "passage", "document", "value"]
# Candidate field names for the source/filename
_SOURCE_FIELDS = ["source", "filename", "file", "url", "title", "name", "origin", "path", "document_id", "doc_id"]
# Candidate field names for the chunk index
_INDEX_FIELDS  = ["chunk_index", "chunk_id", "index", "seq", "sequence", "order", "position", "page"]


def _resolve_payload(payload: dict) -> dict:
    """
    Given a raw Qdrant point payload (arbitrary schema), return a normalised
    dict with keys: text, source, chunk_index.

    Strategy for `text`:
      - Try each name in _TEXT_FIELDS at the top level.
      - If still empty, try one level deep inside a "metadata" sub-dict.
      - Last resort: JSON-dump the whole payload so the content is never lost.

    Strategy for `source` / `chunk_index`:
      - Try candidate names at top level, then inside "metadata".
      - Fall back to sensible defaults (empty string / None).
    """
    if not payload:
        return {"text": "", "source": "", "chunk_index": None}

    meta = payload.get("metadata") or {}
    if not isinstance(meta, dict):
        meta = {}

    def _first(keys, *dicts):
        for d in dicts:
            for k in keys:
                v = d.get(k)
                if v is not None and str(v).strip():
                    return str(v).strip()
        return None

    text = _first(_TEXT_FIELDS, payload, meta)
    if not text:
        # Last resort: dump everything except the metadata sub-dict to avoid duplication
        import json as _j
        flat = {k: v for k, v in payload.items() if k != "metadata" and not isinstance(v, dict)}
        text = _j.dumps(flat, ensure_ascii=False, indent=2) if flat else _j.dumps(payload, ensure_ascii=False, indent=2)

    source = _first(_SOURCE_FIELDS, payload, meta) or ""
    # For source, also try "Bezeichnung" or "BA-Nummer" style keys common in
    # German railway/property datasets — check top-level string values as label
    if not source:
        for k, v in {**payload, **meta}.items():
            if isinstance(v, str) and 3 < len(v) < 80 and k.lower() not in ("content", "text", "body"):
                source = f"{k}: {v}"
                break

    idx = _first(_INDEX_FIELDS, payload, meta)
    chunk_index = int(idx) if idx is not None and str(idx).isdigit() else None

    # ── New structured-metadata fields ──────────────────────────
    # These are present for chapter_aware / csv_rows / json_objects strategies.
    def _str_field(key: str):
        v = payload.get(key) or meta.get(key)
        return str(v).strip() if v is not None else ""

    page          = _str_field("page")
    section_title = _str_field("section_title")
    breadcrumb    = _str_field("breadcrumb")
    csv_headers   = payload.get("csv_headers") or meta.get("csv_headers") or []
    json_key      = payload.get("json_key")   or meta.get("json_key")

    result = {
        "text": text,
        "source": source,
        "chunk_index": chunk_index,
        # structured extras (empty/None when not applicable)
        "page":          page,
        "section_title": section_title,
        "breadcrumb":    breadcrumb,
        "csv_headers":   csv_headers,
        "json_key":      json_key,
    }

    # ── Pass through any other payload fields not already captured ──
    # New/unknown ingestion strategies may add fields we don't explicitly
    # know about (e.g. ingested_at, pdf_page_index, chunk_strategy) — don't
    # silently drop them. Top-level payload keys take priority over
    # same-named keys nested under "metadata".
    _KNOWN_KEYS = {
        "metadata", *_TEXT_FIELDS, *_SOURCE_FIELDS, *_INDEX_FIELDS,
        "page", "section_title", "breadcrumb", "csv_headers", "json_key",
    }
    for k, v in meta.items():
        if k not in _KNOWN_KEYS and k not in result:
            result[k] = v
    for k, v in payload.items():
        if k not in _KNOWN_KEYS:
            result[k] = v

    return result


rag_router = APIRouter(prefix="/api/rag", tags=["rag"])

# ── Helpers ───────────────────────────────────────────────────

def _get_qdrant(url: str) -> "QdrantClient":
    """Return a sync QdrantClient for the given URL."""
    if not _QDRANT_OK:
        raise HTTPException(503, "qdrant-client not installed. Run: pip install qdrant-client")
    return QdrantClient(url=url)


# Qdrant's REST endpoint rejects any single upsert request whose JSON body exceeds
# its configured limit (default 32 MiB, i.e. 33554432 bytes). Large ingests (e.g. a
# multi-hundred-page PDF turned into thousands of chunks) can easily blow past this
# in one shot, so points are always upserted in size-aware sub-batches instead.
_QDRANT_MAX_BATCH_POINTS = 200            # hard cap on points per request
_QDRANT_MAX_BATCH_BYTES  = 24 * 1024 * 1024  # stay comfortably under the 32 MiB limit


def _iter_point_batches(points: list, max_points: int = _QDRANT_MAX_BATCH_POINTS,
                         max_bytes: int = _QDRANT_MAX_BATCH_BYTES):
    """Yield sub-lists of `points`, each kept under both a point-count cap and an
    estimated JSON-payload byte cap, so no single Qdrant upsert request exceeds the
    server's max request size."""
    batch: list = []
    batch_bytes = 0
    for p in points:
        # Rough estimate of this point's serialized size: vector floats + payload JSON.
        est = (len(p.vector) * 12) + len(json.dumps(p.payload, default=str)) + 64
        if batch and (len(batch) >= max_points or batch_bytes + est > max_bytes):
            yield batch
            batch, batch_bytes = [], 0
        batch.append(p)
        batch_bytes += est
    if batch:
        yield batch


def _upsert_points_batched(client: "QdrantClient", collection: str, points: list):
    """Upsert `points` into `collection` across as many requests as needed to stay
    under Qdrant's max payload size. Raises on the first failing batch."""
    for batch in _iter_point_batches(points):
        client.upsert(collection_name=collection, points=batch, wait=True)


async def _get_qdrant_async(url: str) -> "AsyncQdrantClient":
    if not _QDRANT_OK:
        raise HTTPException(503, "qdrant-client not installed. Run: pip install qdrant-client")
    return AsyncQdrantClient(url=url)



# ═══════════════════════════════════════════════════════════════════════════ #
# Chunking engine — adopted from rag-qdrant-app reference implementation     #
# ═══════════════════════════════════════════════════════════════════════════ #
#
# Design principles (from reference app):
#   • _pack() is the key primitive: greedy whole-unit packing (sentences or
#     paragraphs) with carry-back overlap that never splits mid-word/sentence.
#   • _split_recursive: paragraph-level grouping with sentence-level fallback
#     for oversized paragraphs — much better for prose-heavy technical docs.
#   • All size units are CHARACTERS (not words).
#   • PDF: per-page parsing with citation metadata extraction
#     (document_number, page_label, section_title) carried into every chunk.
#   • Records strategy: one chunk per CSV row / JSON object / XLSX row,
#     with every scalar field promoted to Qdrant payload metadata.
#   • "whole" strategy: entire document as one chunk (useful for short files).

_SCALAR_TYPES = (str, int, float, bool, type(None))

# Regex patterns for PDF per-page citation metadata extraction.
# document_number covers: 804.8001, 853.1002, 804.8001A02, 853.8001.V01
_DOC_NUMBER_RE = re.compile(
    r'\b(\d{3}\.\d{3,4}(?:\.\d{2}|[A-Z]\d{2}|\.[A-Z]\d{2})?)\b'
)
_PAGE_LABEL_RE = re.compile(r'Seite\s+(\d+)', re.IGNORECASE)
# Numbered section headings starting with a capital German letter
_HEADING_RE    = re.compile(r'^(\d{1,2})\s+([A-ZÄÖÜ][^\n]{2,80})$', re.MULTILINE)

# Fields copied from a PDF page record into every chunk that came from it
_CITATION_FIELDS = ('document_number', 'page_label', 'pdf_page_index', 'section_title')


# ── Low-level splitters ───────────────────────────────────────────────────────

def _split_fixed(text: str, size: int, overlap: int) -> list[str]:
    """
    Sliding char-window that always cuts at a word boundary.

    Within the last 15% of each window we search backward for the nearest
    whitespace so we never slice mid-word.  If no whitespace is found
    (very long token) we fall back to the hard limit.
    """
    if size <= 0:
        return [text] if text.strip() else []
    step = max(size - overlap, 1)
    chunks: list[str] = []
    start = 0
    while start < len(text):
        end = start + size
        if end < len(text):
            # Search backward for whitespace within the last 15% of the window
            search_from = end - max(1, size // 7)
            ws_pos = text.rfind(' ', search_from, end)
            if ws_pos == -1:
                ws_pos = text.rfind('\n', search_from, end)
            if ws_pos > start:
                end = ws_pos  # cut at word boundary
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        # Advance by step, but align to word boundary too
        next_start = start + step
        if next_start < len(text) and text[next_start] not in (' ', '\n', '\t'):
            # Move to next word start
            ws_pos2 = text.find(' ', next_start)
            if ws_pos2 != -1 and ws_pos2 - next_start < size // 10:
                next_start = ws_pos2 + 1
        start = next_start
        if start >= len(text):
            break
    return chunks


def _pack(units: list[str], size: int, overlap: int, joiner: str) -> list[str]:
    """
    Greedy whole-unit packer with carry-back overlap.

    Packs whole units (sentences or paragraphs) into chunks of at most *size*
    chars. When a chunk is full, the tail units that fit within *overlap* chars
    are carried forward into the next chunk so context survives boundaries
    without duplicating partial words or sentences.
    """
    units = [u.strip() for u in units if u.strip()]
    chunks: list[str] = []
    current: list[str] = []
    current_len = 0
    for unit in units:
        unit_len = len(unit) + len(joiner)
        if current and current_len + unit_len > size:
            chunks.append(joiner.join(current).strip())
            if overlap > 0:
                carry: list[str] = []
                carry_len = 0
                for prev in reversed(current):
                    if carry_len + len(prev) > overlap:
                        break
                    carry.insert(0, prev)
                    carry_len += len(prev) + len(joiner)
                current = carry + [unit]
                current_len = sum(len(c) for c in current) + len(joiner) * len(current)
            else:
                current = [unit]
                current_len = unit_len
        else:
            current.append(unit)
            current_len += unit_len
    if current:
        joined = joiner.join(current).strip()
        if joined:
            chunks.append(joined)
    if chunks:
        return chunks
    fallback = joiner.join(units).strip()
    return [fallback] if fallback else []


def _split_sentence(text: str, size: int, overlap: int) -> list[str]:
    """Pack whole sentences into chunks using _pack()."""
    sentences = re.split(r'(?<=[.!?])\s+', text)
    return _pack(sentences, size, overlap, joiner=' ')


def _split_recursive(text: str, size: int, overlap: int) -> list[str]:
    """
    Paragraph-aware splitting with sentence-level fallback.

    Groups whole paragraphs into chunks up to *size* chars. Paragraphs that
    are themselves larger than *size* are broken at sentence boundaries first.
    Final packing with _pack() ensures whole-unit carry-back overlap.
    """
    paragraphs = [p.strip() for p in re.split(r'\n\s*\n', text) if p.strip()]
    units: list[str] = []
    for p in paragraphs:
        if len(p) <= size:
            units.append(p)
        else:
            units.extend(_split_sentence(p, size, overlap))
    return _pack(units, size, overlap, joiner='\n\n')


def _chunk_text(
    text:     str,
    chunk_size: int = 2000,
    overlap:    int = 200,
    strategy:   str = 'recursive',
) -> list[str]:
    """Dispatch to the selected plain-text chunking strategy."""
    text = text.strip()
    if not text:
        return []
    s = (strategy or 'recursive').lower()
    if s == 'fixed':
        return _split_fixed(text, chunk_size, overlap)
    if s == 'sentence':
        return _split_sentence(text, chunk_size, overlap)
    return _split_recursive(text, chunk_size, overlap)   # default + 'recursive'


# ── PDF per-page parsing with citation metadata ───────────────────────────────

def _detect_pdf_page_meta(
    page_text: str,
    prev_doc_number: 'str | None',
    prev_section_title: 'str | None',
) -> 'tuple[str | None, str | None, str | None]':
    """
    Extract citation metadata from a single PDF page's raw text.

    Used by the whole/chapter_aware/size-based PDF strategies to carry a
    document/page/section citation forward into every chunk. Note: the
    document_number pattern below (NNN.MMMM[A/V]NN) and the German "Seite N"
    page-label pattern were originally tuned for a specific numbered
    document convention — for PDFs that don't match this pattern,
    document_number/page_label will simply come back None/carried-forward,
    which is harmless (chunks are still produced normally either way).

    document_number: NNN.MMMM or NNN.MMMM[A/V]NN
    page_label:      "Seite N" → "N"
    section_title:   last numbered heading on the page (carried forward if absent)
    """
    doc_m = _DOC_NUMBER_RE.search(page_text)
    document_number = doc_m.group(1) if doc_m else prev_doc_number

    page_m = _PAGE_LABEL_RE.search(page_text)
    page_label = page_m.group(1) if page_m else None

    heading_matches = list(_HEADING_RE.finditer(page_text))
    if heading_matches:
        last = heading_matches[-1]
        section_title = f"{last.group(1)} {last.group(2).strip()}"
    else:
        section_title = prev_section_title   # carry forward across pages

    return document_number, page_label, section_title


def _parse_pdf_pages(raw: bytes, filename: str) -> list[dict]:
    """
    Extract text page-by-page from a PDF, attaching citation metadata to each.

    Returns a list of dicts: {text, pdf_page_index, page_label,
    document_number, section_title}.  Uses pypdf for extraction.
    """
    pages: list[dict] = []
    doc_number: 'str | None' = None
    section_title: 'str | None' = None

    # ── pypdf ───────────────────────────────────────────────────────
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(raw))
        for i, page in enumerate(reader.pages):
            text = (page.extract_text() or '').strip()
            if not text:
                continue
            doc_number, page_label, section_title = _detect_pdf_page_meta(
                text, doc_number, section_title)
            pages.append({
                'text': text,
                'pdf_page_index': i + 1,
                'page_label':     page_label,
                'document_number': doc_number,
                'section_title':  section_title,
            })
        return pages
    except ImportError:
        pass
    except Exception:
        pass

    # ── Last resort: extract_text from markitdown / plain read ────────────────
    return [{'text': '', 'pdf_page_index': None, 'page_label': None,
              'document_number': None, 'section_title': None}]


# ── Record-oriented parsing: CSV / XLSX / JSON ───────────────────────────────
# "records" strategy — one chunk per row / object, fields promoted to metadata.

def _row_to_text_and_meta(
    row: dict,
    extra: 'dict | None' = None,
) -> 'tuple[str, dict]':
    """
    Render a flat dict as "Key: Value\\n…" text and return (text, meta).

    Scalar fields (str/int/float/bool/None) go into *meta* for Qdrant
    payload filtering.  Nested structures are JSON-serialised for display
    only (too complex for Qdrant exact-match conditions).
    """
    import json as _json
    meta: dict = {}
    lines: list[str] = []
    for k, v in row.items():
        key = str(k).strip() if k is not None else 'field'
        key = key or 'field'
        if isinstance(v, _SCALAR_TYPES):
            meta[key] = v
            display = '' if v is None else v
        else:
            display = _json.dumps(v, ensure_ascii=False)
        lines.append(f'{key}: {display}')
    if extra:
        meta.update(extra)
    return '\n'.join(lines), meta


def _parse_records_csv(raw: bytes) -> 'list[tuple[str, dict]]':
    import csv, io as _io
    text = raw.decode('utf-8', errors='replace')
    # Auto-detect TSV vs CSV
    first_line = text.splitlines()[0] if text.splitlines() else ''
    dialect = 'excel-tab' if '\t' in first_line else 'excel'
    reader = csv.DictReader(_io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise HTTPException(422, 'CSV has no header row')
    return [_row_to_text_and_meta(dict(row)) for row in reader]


def _parse_records_xlsx(raw: bytes) -> 'list[tuple[str, dict]]':
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(422, 'openpyxl is not installed — cannot ingest XLSX with records strategy')
    import io as _io
    wb = load_workbook(_io.BytesIO(raw), data_only=True, read_only=True)
    out: list[tuple[str, dict]] = []
    for sheet in wb.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue
        header = [
            str(h).strip() if h is not None else f'col{i+1}'
            for i, h in enumerate(header_row)
        ]
        for row in rows_iter:
            if row is None or all(v is None for v in row):
                continue
            row_dict = {
                (header[i] if i < len(header) else f'col{i+1}'): v
                for i, v in enumerate(row)
            }
            out.append(_row_to_text_and_meta(row_dict, extra={'_sheet': sheet.title}))
    return out


def _parse_records_json(raw: bytes) -> 'list[tuple[str, dict]]':
    import json as _json
    text = raw.decode('utf-8', errors='replace')
    try:
        data = _json.loads(text)
    except _json.JSONDecodeError as exc:
        raise HTTPException(422, f'Invalid JSON: {exc}')

    if isinstance(data, list):
        records = data
    elif isinstance(data, dict):
        # Common API-export shapes: {"items": [...]} / {"data": [...]}
        # If exactly one key holds a non-empty list → use that as records
        list_valued = [v for v in data.values() if isinstance(v, list) and v]
        records = list_valued[0] if len(list_valued) == 1 else [data]
    else:
        records = [data]

    out: list[tuple[str, dict]] = []
    for rec in records:
        if isinstance(rec, dict):
            out.append(_row_to_text_and_meta(rec))
        else:
            text_ = _json.dumps(rec, ensure_ascii=False, indent=2)
            meta  = {'value': rec} if isinstance(rec, _SCALAR_TYPES) else {}
            out.append((text_, meta))
    return out


def _parse_records(filename: str, raw: bytes) -> 'list[tuple[str, dict]]':
    """Dispatch to the correct record parser based on file extension."""
    ext = Path(filename).suffix.lower()
    if ext in ('.csv', '.tsv'):
        return _parse_records_csv(raw)
    if ext in ('.xlsx', '.xls'):
        return _parse_records_xlsx(raw)
    if ext == '.json':
        return _parse_records_json(raw)
    raise HTTPException(422, f"'records' strategy not applicable to '{ext or 'this file type'}' files")




# ═══════════════════════════════════════════════════════════════════════════ #
# Chapter/heading-aware chunking — generalized breadcrumb strategy            #
# ═══════════════════════════════════════════════════════════════════════════ #
#
# Generalized heading-aware chunking that works for any document with a
# repeating heading pattern that marks a natural "keep this together" unit —
# Bible chapters, numbered manual sections, markdown-headed chapters, etc.
#
# Each heading's body becomes ONE chunk (preserving full chapter context)
# unless it exceeds chunk_size, in which case it falls back to
# paragraph/sentence/fixed packing (or is kept whole regardless of size, if
# fallback_strategy == "whole"). Every resulting chunk carries its full
# hierarchical breadcrumb ("Title 1 > Title 2 > Title 4", "12.1 > 12.1.1 …")
# as METADATA ONLY — the chunk text itself stays exactly as written in the
# source, with no injected heading/part labels, since that context is already
# available to the LLM via the chunk's payload attributes.

def _md_heading_level(m: 're.Match') -> int:
    return len(m.group(1))


def _md_heading_title(m: 're.Match') -> str:
    return m.group(2).strip()


def _numbered_heading_level(m: 're.Match') -> int:
    return m.group(1).count('.') + 1


def _numbered_heading_title(m: 're.Match') -> str:
    return f"{m.group(1)} {m.group(2).strip()}"


def _single_level(_m: 're.Match') -> int:
    return 1


def _bible_heading_title(m: 're.Match') -> str:
    return f"{m.group(1).strip()} {m.group(2).strip()}"


def _allcaps_heading_title(m: 're.Match') -> str:
    return m.group(1).strip()


# Each preset: a line-anchored regex plus functions that turn a successful
# match into a (level, title) pair. `level` drives the breadcrumb hierarchy —
# e.g. markdown "###" → level 3, numbered "12.1.1" → level 3 — so a deeper
# heading nests under the nearest shallower one, folder-structure style.
_CHAPTER_HEADING_PRESETS: dict = {
    'markdown_headings': {
        're':    re.compile(r'^(#{1,6})\s+(.+?)\s*$', re.MULTILINE),
        'level': _md_heading_level,
        'title': _md_heading_title,
    },
    'numbered': {
        're':    re.compile(r'^\s*(\d{1,3}(?:\.\d{1,3}){0,5})\s{1,4}(\S[^\n]{2,120})\s*$', re.MULTILINE),
        'level': _numbered_heading_level,
        'title': _numbered_heading_title,
    },
    # Matches lines like "Genesis 1", "1 Corinthians 13", "Psalm 23" — the
    # book name and chapter number are captured separately so they can be
    # promoted to their own `book` / `chapter_number` payload fields. Single
    # level only (Bible books don't nest further).
    'bible_book_chapter': {
        're': re.compile(
            r'^\s*((?:[1-3]\s+)?[A-ZÀ-Ý][a-zà-ÿ]+(?:\s[A-ZÀ-Ý][a-zà-ÿ]+)?)\s+(\d{1,3})\s*$',
            re.MULTILINE,
        ),
        'level':   _single_level,
        'title':   _bible_heading_title,
        'is_bible': True,
    },
    'allcaps': {
        're':    re.compile(r"^\s*([A-ZÄÖÜ][A-ZÄÖÜ '\-]{2,60})\s*$", re.MULTILINE),
        'level': _single_level,
        'title': _allcaps_heading_title,
    },
}

# Priority order tried per line when heading_preset == "auto".
_AUTO_PRESET_ORDER = ['numbered', 'markdown_headings', 'bible_book_chapter', 'allcaps']


def _resolve_chapter_matcher(preset: str, custom_regex: str):
    """
    Return a callable `line -> dict|None` that recognizes heading lines for
    the chosen preset and reports their (level, title, book, chapter_number)
    — the building blocks _split_by_chapters uses to construct hierarchical
    breadcrumbs.
    """
    preset = (preset or 'auto').lower()

    if preset == 'custom':
        if not custom_regex or not custom_regex.strip():
            raise HTTPException(422, "chapter_aware: 'custom' heading preset requires a "
                                      "non-empty chapter_heading_regex.")
        try:
            rx = re.compile(custom_regex, re.MULTILINE)
        except re.error as e:
            raise HTTPException(422, f"chapter_aware: invalid custom regex — {e}")

        def _match_custom(line: str):
            m = rx.match(line)
            if not m:
                return None
            g = next((g for g in m.groups() if g), None) if m.groups() else None
            # Custom patterns are treated as single-level (no depth info to
            # infer hierarchy from), matching the previous behavior.
            return {'level': 1, 'title': (g or line).strip(), 'book': '', 'chapter_number': ''}
        return _match_custom

    if preset == 'auto':
        specs = [(name, _CHAPTER_HEADING_PRESETS[name]) for name in _AUTO_PRESET_ORDER]
    elif preset in _CHAPTER_HEADING_PRESETS:
        specs = [(preset, _CHAPTER_HEADING_PRESETS[preset])]
    else:
        raise HTTPException(422, f"chapter_aware: unknown heading preset '{preset}'.")

    def _match(line: str):
        for _name, spec in specs:
            m = spec['re'].match(line)
            if not m:
                continue
            book = chapter_number = ''
            if spec.get('is_bible'):
                book, chapter_number = m.group(1).strip(), m.group(2).strip()
            return {
                'level': spec['level'](m),
                'title': spec['title'](m),
                'book': book,
                'chapter_number': chapter_number,
            }
        return None
    return _match


def _split_by_chapters(text: str, matcher, breadcrumb_sep: str = ' > ') -> list[dict]:
    """
    Split *text* into chapter/section units using *matcher*, building a
    folder-structure-style hierarchical breadcrumb as headings nest and
    un-nest. A deeper heading (e.g. markdown "###", numbered "12.1.1") nests
    under the nearest still-open shallower heading; a new heading at the same
    or shallower level closes out anything deeper, exactly like navigating
    back up a directory tree before entering a sibling folder.

    Returns a list of dicts: {title, breadcrumb, book, chapter_number, level,
    body}. Text before the first heading (if any) becomes its own chapter
    with an empty title so no content is silently dropped.
    """
    lines = text.split('\n')
    stack: list[tuple[int, str]] = []   # [(level, title), ...] currently open
    segments: list[dict] = []
    cur = {'title': '', 'breadcrumb': '', 'book': '', 'chapter_number': '', 'level': 0, 'body': []}

    def _flush():
        body = '\n'.join(cur['body']).strip()
        if body or cur['title']:
            segments.append({
                'title':          cur['title'],
                'breadcrumb':     cur['breadcrumb'],
                'book':           cur['book'],
                'chapter_number': cur['chapter_number'],
                'level':          cur['level'],
                'body':           body,
            })

    for line in lines:
        m = matcher(line)
        if m:
            _flush()
            level = m['level']
            # Close out any open headings at this level or deeper, then push
            # this one — same rule as descending into a new folder.
            while stack and stack[-1][0] >= level:
                stack.pop()
            stack.append((level, m['title']))
            cur = {
                'title':          m['title'],
                'breadcrumb':     breadcrumb_sep.join(t for _, t in stack),
                'book':           m['book'],
                'chapter_number': m['chapter_number'],
                'level':          level,
                'body':           [],
            }
        else:
            cur['body'].append(line)
    _flush()
    return segments


def _chunk_chapter_aware(
    text:              str,
    chunk_size:        int,
    overlap:           int,
    heading_preset:    str = 'auto',
    custom_regex:      str = '',
    fallback_strategy: str = 'recursive',
    breadcrumb_sep:    str = ' > ',
) -> list[tuple[str, dict]]:
    """
    Split *text* into one chunk per detected chapter/heading, preserving full
    chapter context and a hierarchical breadcrumb. Chapters larger than
    chunk_size fall back to paragraph/sentence/fixed packing (*fallback_strategy*),
    unless fallback_strategy == "whole", in which case oversized chapters are
    kept as a single chunk regardless of size. Chunk text never includes the
    breadcrumb, chapter title, or part label — that context lives entirely in
    the returned metadata (chapter_title, breadcrumb, chunk_part, …).
    """
    matcher = _resolve_chapter_matcher(heading_preset, custom_regex)
    chapters = _split_by_chapters(text, matcher, breadcrumb_sep)

    if not chapters:
        # No headings matched at all — fall back to plain packing over the
        # whole document rather than silently dropping everything.
        plain = _chunk_text(text, chunk_size, overlap,
                             'recursive' if fallback_strategy == 'whole' else fallback_strategy)
        return [(c, {}) for c in plain]

    results: list[tuple[str, dict]] = []
    for idx, ch in enumerate(chapters):
        body = ch['body']
        if not body:
            continue
        breadcrumb = ch['breadcrumb'] or f"Section {idx + 1}"
        base_meta = {
            'chapter_title': ch['title'] or breadcrumb,
            'breadcrumb':    breadcrumb,
            'chapter_index': idx,
            'heading_level': ch['level'],
        }
        if ch['book']:
            base_meta['book'] = ch['book']
        if ch['chapter_number']:
            base_meta['chapter_number'] = ch['chapter_number']

        keep_whole = fallback_strategy == 'whole' or len(body) <= chunk_size
        if keep_whole:
            results.append((body, base_meta))
        else:
            parts = _chunk_text(body, chunk_size, overlap, fallback_strategy)
            n_parts = len(parts)
            for p_i, part in enumerate(parts):
                part_meta = dict(base_meta)
                part_meta['chunk_part'] = p_i + 1
                part_meta['chunk_parts_total'] = n_parts
                results.append((part, part_meta))

    return results


async def _embed_ollama(texts: list[str], model: str, base_url: str) -> list[list[float]]:
    """Fetch embeddings from Ollama /api/embed (batch).
    Falls back to /api/embeddings (single) for older Ollama builds."""
    base = base_url.rstrip("/")
    async with httpx.AsyncClient(timeout=120.0) as client:
        try:
            resp = await client.post(f"{base}/api/embed", json={"model": model, "input": texts})
            resp.raise_for_status()
            data = resp.json()
            if "embeddings" in data:
                return data["embeddings"]
        except Exception:
            pass
        # Fallback: one request per text
        results = []
        for text in texts:
            resp = await client.post(f"{base}/api/embeddings", json={"model": model, "prompt": text})
            resp.raise_for_status()
            results.append(resp.json()["embedding"])
        return results


async def _embed_openai(texts: list[str], model: str, base_url: str) -> list[list[float]]:
    """Fetch embeddings from OpenAI-compatible /v1/embeddings."""
    base = base_url.rstrip("/")
    async with httpx.AsyncClient(timeout=120.0) as client:
        resp = await client.post(
            f"{base}/v1/embeddings",
            json={"model": model, "input": texts},
        )
        resp.raise_for_status()
        data = resp.json()
        return [item["embedding"] for item in data["data"]]


async def _embed(texts: list[str], model: str, base_url: str, flavor: str = "ollama") -> list[list[float]]:
    """Dispatch to the right embedding API based on *flavor*."""
    if flavor == "openai":
        return await _embed_openai(texts, model, base_url)
    return await _embed_ollama(texts, model, base_url)


# ── Plain-text file suffixes (read directly without conversion) ───────────────

_PLAINTEXT_SUFFIXES = {
    ".txt", ".md", ".markdown", ".rst",
    ".json", ".jsonl", ".ndjson",
    ".csv", ".tsv",
    ".yaml", ".yml",
    ".toml", ".ini", ".cfg", ".conf", ".env",
    ".py", ".js", ".ts", ".jsx", ".tsx",
    ".html", ".htm", ".xml", ".svg",
    ".css", ".scss", ".less",
    ".sh", ".bash", ".zsh", ".sql", ".graphql",
    ".log", ".diff", ".patch",
}


def _extract_text(filepath: str, filename: str) -> str:
    """Extract plain text from a file on disk.
    1. Plain-text suffixes -> read directly (markitdown returns empty for JSON/CSV).
    2. PDF / Office / binary -> markitdown.
    3. Fallback -> UTF-8 read.
    """
    import json as _json
    suffix = Path(filename).suffix.lower()

    if suffix in _PLAINTEXT_SUFFIXES:
        try:
            raw = Path(filepath).read_text(encoding="utf-8", errors="replace")
            if suffix in (".json", ".jsonl", ".ndjson"):
                try:
                    if suffix == ".json":
                        obj = _json.loads(raw)
                        return _json.dumps(obj, ensure_ascii=False, indent=2)
                    else:
                        parts = []
                        for line in raw.splitlines():
                            line = line.strip()
                            if not line:
                                continue
                            try:
                                parts.append(_json.dumps(_json.loads(line), ensure_ascii=False, indent=2))
                            except Exception:
                                parts.append(line)
                        return "\n\n".join(parts)
                except Exception:
                    pass
            return raw
        except Exception as e:
            raise HTTPException(500, f"Cannot read file: {e}")

    if _markitdown:
        try:
            result = _markitdown.convert(filepath)
            text = (result.text_content or "").strip()
            if text:
                return text
        except Exception:
            pass

    try:
        return Path(filepath).read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        raise HTTPException(500, f"Cannot read file: {e}")


def _parse_collection_info(info) -> tuple:
    """
    Extract (dim, distance, points_count) from a CollectionInfo object.
    Handles the three shapes qdrant-client has used across versions:
      1. info.config.params.vectors  → VectorParams  (old single-vector)
      2. info.config.params.vectors  → dict[str, VectorParams]  (named vectors)
      3. Direct REST-style dict (fallback)
    Also reads points_count from info.points_count (may be None in newer builds).
    """
    dim = "?"
    dist = "?"
    try:
        vec_cfg = info.config.params.vectors
        if vec_cfg is None:
            pass
        elif hasattr(vec_cfg, "size"):
            # Shape 1 — single anonymous vector config
            dim  = vec_cfg.size
            dist = vec_cfg.distance.value if hasattr(vec_cfg.distance, "value") else str(vec_cfg.distance)
        elif isinstance(vec_cfg, dict) and vec_cfg:
            # Shape 2 — named vectors dict; take the first entry
            first = next(iter(vec_cfg.values()))
            dim  = first.size if hasattr(first, "size") else "?"
            dist = first.distance.value if hasattr(first.distance, "value") else str(getattr(first, "distance", "?"))
        else:
            # Shape 3 — try attribute access on the object itself
            dim  = getattr(vec_cfg, "size", "?")
            dist = getattr(getattr(vec_cfg, "distance", None), "value", str(getattr(vec_cfg, "distance", "?")))
    except Exception:
        pass

    # points_count is None in Qdrant >= 1.9 when payload indexing is deferred;
    # fall back to vectors_count which is always populated.
    pts = info.points_count
    if pts is None:
        pts = getattr(info, "vectors_count", None) or 0
    return dim, dist, int(pts)


@rag_router.get("/collections")
def list_collections(qdrant_url: str = "http://localhost:6333"):
    """List all Qdrant collections with metadata."""
    client = _get_qdrant(qdrant_url)
    try:
        resp = client.get_collections()
        collections = []
        for col in resp.collections:
            try:
                info = client.get_collection(col.name)
                dim, dist, pts = _parse_collection_info(info)
                collections.append({
                    "name":         col.name,
                    "points_count": pts,
                    "dimension":    dim,
                    "distance":     dist,
                    "status":       str(info.status),
                })
            except Exception as ex:
                collections.append({"name": col.name, "error": str(ex)})
        return {"collections": collections}
    except Exception as e:
        raise HTTPException(502, f"Qdrant error: {e}")


class CreateCollectionRequest(BaseModel):
    name: str
    dimension: int = 768
    distance: str = "Cosine"   # Cosine | Dot | Euclid | Manhattan
    qdrant_url: str = "http://localhost:6333"

@rag_router.post("/collections")
def create_collection(req: CreateCollectionRequest):
    """Create a new Qdrant collection."""
    client = _get_qdrant(req.qdrant_url)
    dist_map = {
        "cosine":    Distance.COSINE,
        "dot":       Distance.DOT,
        "euclid":    Distance.EUCLID,
        "manhattan": Distance.MANHATTAN,
    }
    dist = dist_map.get(req.distance.lower(), Distance.COSINE)
    try:
        client.create_collection(
            collection_name=req.name,
            vectors_config=VectorParams(size=req.dimension, distance=dist),
        )
        return {"ok": True, "name": req.name}
    except Exception as e:
        raise HTTPException(409, f"Could not create collection: {e}")


@rag_router.delete("/collections/{name}")
def delete_collection(name: str, qdrant_url: str = "http://localhost:6333"):
    """Delete a Qdrant collection and all its points."""
    client = _get_qdrant(qdrant_url)
    try:
        client.delete_collection(name)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(502, f"Qdrant error: {e}")


@rag_router.get("/collections/{name}/points")
def list_points(name: str, limit: int = 50, offset: int = 0,
                qdrant_url: str = "http://localhost:6333"):
    """Return a page of points (without vectors) so the UI can inspect ingested chunks."""
    client = _get_qdrant(qdrant_url)
    try:
        result = client.scroll(
            collection_name=name,
            limit=limit,
            offset=offset,
            with_payload=True,
            with_vectors=False,
        )
        points, next_offset = result
        return {
            "points": [
                {
                    "id":      str(p.id),
                    "payload": p.payload,
                    # resolved fields for display — works regardless of schema
                    **_resolve_payload(p.payload or {}),
                }
                for p in points
            ],
            "next_offset": str(next_offset) if next_offset else None,
        }
    except Exception as e:
        raise HTTPException(502, f"Qdrant error: {e}")


# ── Field discovery (for Settings → RAG whitelist/blacklist picker) ───────

@rag_router.get("/collections/{name}/fields")
def get_collection_fields(name: str, qdrant_url: str = "http://localhost:6333", sample_size: int = 100):
    """
    Sample up to `sample_size` points from the collection and return the
    union of resolved payload keys, each with how many of the sampled
    points contain it (coverage) and a short example value. Powers the
    field whitelist/blacklist picker in Settings → RAG — lets the user see
    exactly which keys exist before deciding which to include/exclude
    from the rag_search tool's result.
    """
    client = _get_qdrant(qdrant_url)
    try:
        points, _ = client.scroll(
            collection_name=name,
            limit=sample_size,
            with_payload=True,
            with_vectors=False,
        )
    except Exception as e:
        raise HTTPException(502, f"Qdrant error: {e}")

    sampled = len(points)
    field_info: dict[str, dict] = {}
    for p in points:
        resolved = _resolve_payload(p.payload or {})
        for k, v in resolved.items():
            info = field_info.setdefault(k, {"count": 0, "example": None})
            info["count"] += 1
            if info["example"] is None and v not in (None, "", [], {}):
                ex = v
                if isinstance(ex, str) and len(ex) > 100:
                    ex = ex[:100] + "…"
                info["example"] = ex

    fields = [
        {
            "key":      k,
            "count":    info["count"],
            "coverage": round(info["count"] / sampled, 2) if sampled else 0,
            "example":  info["example"],
        }
        for k, info in sorted(field_info.items())
    ]
    return {"sampled": sampled, "fields": fields}


# ── Ingestion endpoint ────────────────────────────────────────

@rag_router.post("/ingest")
async def ingest_file(
    file: UploadFile          = File(...),
    collection: str           = Form(...),
    embed_model: str          = Form("nomic-embed-text"),
    embed_flavor: str         = Form("ollama"),   # "ollama" | "openai"
    ollama_base: str          = Form("http://localhost:11434"),
    qdrant_url: str           = Form("http://localhost:6333"),
    chunk_size: int           = Form(2000),
    chunk_overlap: int        = Form(200),
    chunk_strategy: str       = Form("recursive"),
    # "csv_rows" → "records", "json_objects" → "records"
    chapter_heading_preset: str = Form("auto"),
    # "auto" | "markdown_headings" | "numbered" | "bible_book_chapter"
    # | "allcaps" | "custom" — only used by the chapter_aware strategy.
    chapter_heading_regex: str  = Form(""),
    # User-supplied regex, only required when chapter_heading_preset == "custom".
    chapter_fallback: str       = Form("recursive"),
    # "recursive" | "sentence" | "fixed" — packing method used only for
    # chapters that exceed chunk_size. "whole" keeps every chapter as a
    # single chunk regardless of size (never splits, even if oversized).
    chapter_breadcrumb_sep: str = Form(" > "),
    # Separator joining hierarchy levels in the breadcrumb, e.g. " > ", " / ",
    # " -> ". Only affects metadata (breadcrumb text is never injected into
    # the chunk body itself).
):
    """
    Extract text from an uploaded file, chunk it, embed each chunk and upsert
    into Qdrant.  Returns Server-Sent Events so the UI can show live progress.

    Strategy overview
    -----------------
    fixed / sentence / recursive
        Size-based char-window strategies.  PDFs are parsed page-by-page and
        each page's citation metadata (document_number, page_label,
        section_title) is carried into every chunk from that page.

    records  (formerly csv_rows / json_objects)
        One chunk per row/object.  Every scalar field is promoted to a Qdrant
        payload metadata key so the LLM can filter or cite it exactly.
        Supports CSV, TSV, JSON (array or API-export dict), XLSX.

    whole
        Entire document text as a single chunk.  Useful for short files or
        when downstream retrieval should always return the full document.

    chapter_aware
        Generalized heading-aware chunking: splits on a configurable heading
        pattern (chapter_heading_preset) and keeps each chapter's full text as
        ONE chunk — no fixed-length slicing — so context never gets fragmented
        mid-chapter. Nested headings (markdown "#"/"##"/"###", numbered
        "12"/"12.1"/"12.1.1") build a folder-structure breadcrumb, e.g.
        "Title 1 > Title 2 > Title 4" or "12 Title > 12.1 Title". Chapters
        exceeding chunk_size fall back to chapter_fallback (paragraph/
        sentence/fixed) packing — or, with fallback "whole", are never split
        at all. The breadcrumb/chapter title/part number live ONLY in chunk
        metadata (chapter_title, breadcrumb, chunk_part, …); the chunk text
        itself is never prefixed with them, since that's redundant once the
        RAG layer hands the attributes to the LLM separately. Good fit for
        Bibles, books, and numbered manuals where "which chapter is this
        from" must survive retrieval.
    """
    filename = file.filename or "unknown"
    suffix   = Path(filename).suffix.lower()
    raw      = await file.read()

    async def _stream():
        def _event(data: dict) -> str:
            return f"data: {json.dumps(data)}\n\n"

        # Normalise legacy aliases
        strategy = (chunk_strategy or "recursive").lower()
        if strategy in ("csv_rows", "json_objects"):
            strategy = "records"

        # ── chunk_specs: list of (text, extra_payload_dict) ─────────────────
        chunk_specs: list[tuple[str, dict]] = []

        # ── records strategy ─────────────────────────────────────────────────
        if strategy == "records":
            yield _event({"stage": "extracting", "message": f"Parsing records from {filename}…"})
            try:
                records = _parse_records(filename, raw)
            except HTTPException as e:
                yield _event({"stage": "error", "message": e.detail}); return
            except Exception as e:
                yield _event({"stage": "error", "message": f"Parse error: {e}"}); return

            if not records:
                yield _event({"stage": "error", "message": "No records found in file."}); return

            yield _event({"stage": "chunking",
                          "message": f"{len(records)} records found.", "total": len(records)})
            chunk_specs = [(text, meta) for text, meta in records]

        # ── whole strategy ───────────────────────────────────────────────────
        elif strategy == "whole":
            yield _event({"stage": "extracting", "message": f"Extracting text from {filename}…"})
            pages = _parse_pdf_pages(raw, filename) if suffix == ".pdf" else None

            if pages is None:
                # Non-PDF: use _extract_text
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp.write(raw); tmp_path = tmp.name
                try:
                    text = _extract_text(tmp_path, filename)
                except HTTPException as e:
                    yield _event({"stage": "error", "message": e.detail}); return
                finally:
                    try: os.unlink(tmp_path)
                    except Exception: pass
                chunk_specs = [(text.strip(), {})] if text.strip() else []
            else:
                full_text = "\n\n".join(p["text"] for p in pages if p["text"])
                # Aggregate citation metadata across all pages
                agg: dict = {}
                doc_nums = sorted({p["document_number"] for p in pages if p.get("document_number")})
                pg_labels = [p["page_label"] for p in pages if p.get("page_label")]
                if doc_nums:
                    agg["document_number"] = doc_nums[0] if len(doc_nums) == 1 else ", ".join(doc_nums)
                if pg_labels:
                    agg["page_label"] = pg_labels[0] if len(pg_labels) == 1 else f"{pg_labels[0]}-{pg_labels[-1]}"
                chunk_specs = [(full_text, agg)] if full_text.strip() else []

            if not chunk_specs:
                yield _event({"stage": "error", "message": "No text extracted."}); return
            yield _event({"stage": "chunking", "message": "1 chunk (whole document).", "total": 1})

        # ── chapter_aware strategy ─────────────────────────────────────────────
        elif strategy == "chapter_aware":
            yield _event({"stage": "extracting", "message": f"Extracting text from {filename}…"})

            if suffix == ".pdf":
                pages = _parse_pdf_pages(raw, filename)
                if not pages:
                    yield _event({"stage": "error", "message": "No extractable text in PDF."}); return
                text = "\n\n".join(p["text"] for p in pages if p["text"])
            else:
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp.write(raw); tmp_path = tmp.name
                try:
                    text = _extract_text(tmp_path, filename)
                except HTTPException as e:
                    yield _event({"stage": "error", "message": e.detail}); return
                finally:
                    try: os.unlink(tmp_path)
                    except Exception: pass

            if not text.strip():
                yield _event({"stage": "error", "message": "No text extracted."}); return

            yield _event({"stage": "chunking",
                          "message": f"Splitting '{filename}' by chapters/headings "
                                     f"('{chapter_heading_preset}' pattern)…"})
            try:
                chunk_specs = _chunk_chapter_aware(
                    text, chunk_size, chunk_overlap,
                    heading_preset=chapter_heading_preset,
                    custom_regex=chapter_heading_regex,
                    fallback_strategy=chapter_fallback,
                    breadcrumb_sep=chapter_breadcrumb_sep,
                )
            except HTTPException as e:
                yield _event({"stage": "error", "message": e.detail}); return

            if not chunk_specs:
                yield _event({"stage": "error", "message": "No chunks produced."}); return

            n_chapters = len({m["chapter_index"] for _, m in chunk_specs if "chapter_index" in m}) \
                or len(chunk_specs)
            yield _event({"stage": "chunking",
                          "message": f"\u2713 {len(chunk_specs)} chunk(s) from {n_chapters} "
                                     f"chapter(s)/section(s) in {filename}.",
                          "total": len(chunk_specs)})

        # ── size-based strategies: fixed / sentence / recursive ──────────────
        else:
            yield _event({"stage": "extracting", "message": f"Extracting text from {filename}…"})

            # PDFs: parse page-by-page to carry citation metadata per chunk
            if suffix == ".pdf":
                pages = _parse_pdf_pages(raw, filename)
                if not pages:
                    yield _event({"stage": "error", "message": "No extractable text in PDF."}); return
                n_pages = len(pages)
                yield _event({"stage": "chunking",
                              "message": f"Chunking {n_pages} pages with '{strategy}' strategy…",
                              "total": n_pages})
                for i, page in enumerate(pages):
                    page_meta = {k: page[k] for k in _CITATION_FIELDS if page.get(k) is not None}
                    for sub in _chunk_text(page["text"], chunk_size, chunk_overlap, strategy):
                        chunk_specs.append((sub, page_meta))
                    if i < 10 or (i + 1) % 10 == 0 or i == n_pages - 1:
                        yield _event({
                            "stage":   "chunking",
                            "message": (
                                f"Page {i+1}/{n_pages} — "
                                f"{len(chunk_specs)} chunks so far"
                            ),
                            "done":  i + 1,
                            "total": n_pages,
                        })
            else:
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp.write(raw); tmp_path = tmp.name
                try:
                    text = _extract_text(tmp_path, filename)
                except HTTPException as e:
                    yield _event({"stage": "error", "message": e.detail}); return
                finally:
                    try: os.unlink(tmp_path)
                    except Exception: pass

                if not text.strip():
                    yield _event({"stage": "error", "message": "No text extracted."}); return
                for sub in _chunk_text(text, chunk_size, chunk_overlap, strategy):
                    chunk_specs.append((sub, {}))

            total_chunks = len(chunk_specs)
            yield _event({"stage": "chunking",
                          "message": f"{total_chunks} chunks created.", "total": total_chunks})
            if not total_chunks:
                yield _event({"stage": "error", "message": "No chunks produced."}); return

        total = len(chunk_specs)

        chunk_texts = [t for t, _ in chunk_specs]
        chunk_metas = [m for _, m in chunk_specs]

        # ── Embed in batches of 32 ───────────────────────────────────────────
        BATCH = 32
        vectors: list[list[float]] = []
        for batch_start in range(0, total, BATCH):
            batch = chunk_texts[batch_start : batch_start + BATCH]
            yield _event({
                "stage":   "embedding",
                "message": f"Embedding chunks {batch_start + 1}–{min(batch_start + BATCH, total)} / {total}…",
                "done":    batch_start,
                "total":   total,
            })
            try:
                batch_vecs = await _embed(batch, embed_model, ollama_base, embed_flavor)
                vectors.extend(batch_vecs)
            except Exception as e:
                yield _event({"stage": "error", "message": f"Embedding error: {e}"}); return

        # ── Upsert into Qdrant ───────────────────────────────────────────────
        yield _event({"stage": "upserting", "message": f"Upserting {total} points into Qdrant…"})
        try:
            client = _get_qdrant(qdrant_url)
            ingested_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
            points = [
                PointStruct(
                    id=str(uuid.uuid4()),
                    vector=vec,
                    payload={
                        "text":           chunk,
                        "source":         filename,
                        "chunk_index":    i,
                        "chunk_strategy": strategy,
                        "ingested_at":    ingested_at,
                        **chunk_metas[i],
                    },
                )
                for i, (chunk, vec) in enumerate(zip(chunk_texts, vectors))
            ]
            # Qdrant caps request body size (default 32 MiB), so upsert in
            # size-aware sub-batches rather than one giant request.
            upserted = 0
            for batch in _iter_point_batches(points):
                client.upsert(collection_name=collection, points=batch, wait=True)
                upserted += len(batch)
                yield _event({
                    "stage":   "upserting",
                    "message": f"Upserting into Qdrant… {upserted}/{total} points",
                    "done":    upserted,
                    "total":   total,
                })
        except Exception as e:
            yield _event({"stage": "error", "message": f"Qdrant upsert error: {e}"}); return

        yield _event({
            "stage":      "done",
            "message":    f"\u2713 {total} chunks from '{filename}' added to '{collection}'.",
            "chunks":     total,
            "collection": collection,
        })

    return StreamingResponse(
        _stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── Search endpoint ───────────────────────────────────────────

class SearchRequest_(BaseModel):
    collection: str
    query: str
    top_k: int = 5
    score_threshold: float = 0.0
    embed_model: str = "nomic-embed-text"
    embed_flavor: str = "ollama"
    ollama_base: str = "http://localhost:11434"
    qdrant_url: str = "http://localhost:6333"

@rag_router.post("/search")
async def search(req: SearchRequest_):
    """Embed *query* and return top-k matching chunks from the collection.
    Uses query_points() — the unified API available in qdrant-client >= 1.7.
    The old .search() method was removed in qdrant-client 1.13+.
    """
    try:
        vecs = await _embed([req.query], req.embed_model, req.ollama_base, req.embed_flavor)
    except Exception as e:
        raise HTTPException(502, f"Embedding error: {e}")

    query_vec = vecs[0]
    client = _get_qdrant(req.qdrant_url)
    try:
        # query_points replaces the removed .search() in qdrant-client >= 1.7
        response = client.query_points(
            collection_name=req.collection,
            query=query_vec,
            limit=req.top_k,
            score_threshold=req.score_threshold if req.score_threshold > 0 else None,
            with_payload=True,
        )
        # response.points is a list of ScoredPoint
        hits = response.points
    except Exception as e:
        raise HTTPException(502, f"Qdrant search error: {e}")

    return {
        "results": [
            {
                "id":    str(r.id),
                "score": round(r.score, 6),
                **_resolve_payload(r.payload or {}),
            }
            for r in hits
        ]
    }


# ── Module-level search helper (importable by server.py) ────────
async def rag_search_chunks(
    query: str,
    collection: str,
    embed_model: str,
    embed_flavor: str,
    ollama_base: str,
    qdrant_url: str,
    top_k: int = 5,
) -> list[dict]:
    """
    Embed *query* and return top-k matching chunks.
    Called by server.py's chat-generation jobs to inject RAG context
    into the system prompt before the model sees the user message.
    Returns a list of dicts: [{text, source, chunk_index, score}, ...]
    """
    if not _QDRANT_OK:
        return []
    try:
        vecs = await _embed([query], embed_model, ollama_base, embed_flavor)
        query_vec = vecs[0]
        client = _get_qdrant(qdrant_url)
        response = client.query_points(
            collection_name=collection,
            query=query_vec,
            limit=top_k,
            with_payload=True,
        )
        return [
            {
                **_resolve_payload(r.payload or {}),
                "score": round(r.score, 4),
            }
            for r in response.points
        ]
    except Exception:
        return []


# ── Embedding model listing ───────────────────────────────────

@rag_router.get("/embed-models")
async def list_embed_models(ollama_base: str = "http://localhost:11434", flavor: str = "ollama"):
    """
    Return available embedding models from Ollama or OpenAI-compatible API.

    For Ollama, filters the full model list down to known embedding models by
    checking the model name against common embed-model keywords.  This avoids
    showing large language models (llama3, mistral, …) in the embed dropdown.
    """
    # Keywords that appear in embedding model names but NOT in plain LLMs.
    # Checked as case-insensitive substrings of the model name.
    _EMBED_KEYWORDS = [
        "embed", "embedding",          # nomic-embed-text, mxbai-embed-large, …
        "minilm", "bge-",              # sentence-transformers family
        "e5-",                          # e5-mistral, e5-small, …
        "gte-",                         # gte-base, gte-large
        "instructor",                   # instructor-xl
        "all-minilm",
        "paraphrase",
        "multilingual",
        "sentence",
    ]

    def _is_embed(name: str) -> bool:
        n = name.lower()
        return any(kw in n for kw in _EMBED_KEYWORDS)

    base = ollama_base.rstrip("/")
    async with httpx.AsyncClient(timeout=10.0) as client:
        try:
            if flavor == "openai":
                resp = await client.get(f"{base}/v1/models")
                resp.raise_for_status()
                all_models = [m["id"] for m in resp.json().get("data", [])]
            else:
                resp = await client.get(f"{base}/api/tags")
                resp.raise_for_status()
                all_models = [m["name"] for m in resp.json().get("models", [])]

            # Filter to embed models; fall back to full list if nothing matched
            # (e.g. user has only one model and it has a non-standard name)
            filtered = [m for m in all_models if _is_embed(m)]
            return {"models": filtered if filtered else all_models}
        except Exception as e:
            return {"models": [], "error": str(e)}


@rag_router.get("/probe-dimension")
async def probe_dimension(
    model:      str = "nomic-embed-text",
    ollama_base:str = "http://localhost:11434",
    flavor:     str = "ollama",
):
    """
    Embed a short test string and return the vector dimension.
    Used by the UI "Probe" button so the user doesn't have to guess the
    correct dimension when creating a new collection.
    """
    try:
        vecs = await _embed(["hello"], model, ollama_base, flavor)
        if not vecs or not vecs[0]:
            raise ValueError("Empty embedding returned")
        return {"dimension": len(vecs[0]), "model": model}
    except Exception as e:
        raise HTTPException(502, f"Could not probe dimension: {e}")
